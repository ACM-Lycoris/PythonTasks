#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ACM 队伍数据管理平台 — Flask + Pandas 全栈应用
CSV/Excel 导入导出 | 表格浏览 | 内联编辑 | 批量处理 | 数据类型识别
可视化图表 | 撤销重做 | 行操作 | 全局搜索 | 条件格式 | 列统计
启动: python app.py  →  http://127.0.0.1:5000
"""

import io, re, os, json, threading
import pandas as pd
import numpy as np
from collections import deque
from flask import Flask, request, jsonify, send_file, render_template_string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------------------------------------------------------------
app = Flask(__name__)
app.secret_key = "acm-neon-datatable"
DATA_LOCK = threading.Lock()

# ---------------------------------------------------------------------------
# 撤销/重做 历史栈 (最多 50 层)
# ---------------------------------------------------------------------------
UNDO_STACK: deque = deque(maxlen=50)
REDO_STACK: deque = deque(maxlen=50)


def _push_history():
    """将当前 DF 深拷贝压入撤销栈"""
    global UNDO_STACK
    UNDO_STACK.append(DF.copy())


def _clear_redo():
    REDO_STACK.clear()


# ---------------------------------------------------------------------------
# 默认数据 — 3 支 ACM 队伍
# ---------------------------------------------------------------------------
DEFAULT_CSV = """队伍,姓名,年级,专业,AC数,总提交数,AC率
代码敲不队,张三,大三,计科,145,230,63.0%
代码敲不队,李四,大二,计科,87,156,55.8%
代码敲不队,王五,大三,计科,203,298,68.1%
深夜AC自动机,赵六,大四,计科,312,420,74.3%
深夜AC自动机,钱七,大三,计科,178,245,72.7%
深夜AC自动机,孙八,大二,计科,56,89,62.9%
边界条件不存在的,周九,大三,计科,267,345,77.4%
边界条件不存在的,吴十,大二,计科,98,167,58.7%
边界条件不存在的,郑十一,大四,计科,189,234,80.8%"""


def _load_default() -> pd.DataFrame:
    return pd.read_csv(io.StringIO(DEFAULT_CSV), dtype=str).fillna("")


DF: pd.DataFrame = _load_default()


def _reset_default():
    global DF
    DF = _load_default()


# ---------------------------------------------------------------------------
# 辅助
# ---------------------------------------------------------------------------
_ZH_RE = re.compile(r"[一-鿿㐀-䶿豈-﫿]")


def _is_chinese(ch: str) -> bool:
    return bool(_ZH_RE.match(ch))


def _smart_sort(df: pd.DataFrame, col: str, ascending: bool = True) -> pd.DataFrame:
    """智能排序：若列中大部分值可转为数值，则按数值排序；否则按字符串排序。"""
    numeric_vals = pd.to_numeric(df[col], errors="coerce")
    if numeric_vals.notna().sum() >= len(df) * 0.5:
        return (df.assign(__sort_key__=numeric_vals)
                  .sort_values(by="__sort_key__", ascending=ascending)
                  .drop(columns=["__sort_key__"])
                  .reset_index(drop=True))
    return df.sort_values(by=col, ascending=ascending).reset_index(drop=True)


def df_to_csv_bytes(df: pd.DataFrame) -> io.BytesIO:
    buf = io.BytesIO()
    buf.write(b"\xef\xbb\xbf")
    buf.write(df.to_csv(index=False, lineterminator="\n").encode("utf-8"))
    buf.seek(0)
    return buf


def _detect_type(series: pd.Series) -> str:
    """检测列数据类型"""
    try:
        pd.to_numeric(series)
        return "数值"
    except (ValueError, TypeError):
        return "文本"


# ---------------------------------------------------------------------------
# 前端 — 暗色科技风 + 全功能面板
# ---------------------------------------------------------------------------
INDEX_HTML = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>ACM 队伍数据管理平台</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.7/dist/chart.umd.min.js">
</script>
<style>
  :root {
    --bg: #0a0e17;
    --card: #111827;
    --card2: #161e2c;
    --text: #e2e8f0;
    --text2: #94a3b8;
    --cyan: #00e5ff;
    --pink: #ff4081;
    --green: #00e676;
    --amber: #ffab00;
    --red: #ff1744;
    --border: #1e2d45;
    --radius: 8px;
    --glow-cyan: 0 0 20px rgba(0,229,255,0.15);
    --glow-pink: 0 0 20px rgba(255,64,129,0.15);
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: "Inter","Segoe UI","Microsoft YaHei",sans-serif;
    background: var(--bg);
    color: var(--text);
    min-height: 100vh;
    overflow-x: hidden;
  }
  body::before {
    content: "";
    position: fixed; inset: 0; z-index: 0; pointer-events: none;
    background-image:
      linear-gradient(rgba(0,229,255,0.03) 1px, transparent 1px),
      linear-gradient(90deg, rgba(0,229,255,0.03) 1px, transparent 1px);
    background-size: 48px 48px;
    animation: gridPulse 8s ease-in-out infinite;
  }
  @keyframes gridPulse {
    0%,100% { opacity: 0.6; }
    50% { opacity: 1; }
  }
  .app { position: relative; z-index: 1; max-width: 1600px; margin: 0 auto; padding: 16px; }

  header {
    display: flex; gap: 10px; align-items: center; flex-wrap: wrap;
    margin-bottom: 14px; padding: 14px 20px;
    background: var(--card); border: 1px solid var(--border);
    border-radius: 12px; box-shadow: var(--glow-cyan);
  }
  header h1 {
    font-size: 22px; font-weight: 700; margin-right: auto;
    background: linear-gradient(135deg, var(--cyan), #80deea);
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    background-clip: text; letter-spacing: 1px;
  }
  .btn {
    padding: 8px 15px; border: 1px solid transparent; border-radius: var(--radius);
    cursor: pointer; font-size: 12px; font-weight: 600; letter-spacing: 0.3px;
    transition: all 0.2s ease; position: relative; overflow: hidden;
    white-space: nowrap;
  }
  .btn::after {
    content: ""; position: absolute; inset: 0;
    background: linear-gradient(135deg, transparent, rgba(255,255,255,0.06));
    opacity: 0; transition: opacity 0.2s;
  }
  .btn:hover::after { opacity: 1; }
  .btn:hover { transform: translateY(-2px); }
  .btn-primary { background: var(--cyan); color: #000; border-color: var(--cyan); box-shadow: 0 0 16px rgba(0,229,255,0.25); }
  .btn-success { background: var(--green); color: #000; border-color: var(--green); box-shadow: 0 0 16px rgba(0,230,118,0.25); }
  .btn-danger  { background: var(--red); color: #fff; border-color: var(--red); box-shadow: 0 0 16px rgba(255,23,68,0.25); }
  .btn-warn   { background: var(--amber); color: #000; border-color: var(--amber); box-shadow: 0 0 16px rgba(255,171,0,0.25); }
  .btn-outline {
    background: transparent; color: var(--cyan); border: 1px solid var(--cyan);
    box-shadow: 0 0 10px rgba(0,229,255,0.08);
  }
  .btn-outline:hover { background: rgba(0,229,255,0.08); }
  .btn-sm {
    padding: 5px 11px; font-size: 11px; font-weight: 600;
    border: 1px solid var(--cyan); border-radius: 5px; cursor: pointer;
    background: transparent; color: var(--cyan);
    white-space: nowrap; transition: all 0.2s;
    text-transform: uppercase; letter-spacing: 0.5px;
  }
  .btn-sm:hover { background: var(--cyan); color: #000; box-shadow: 0 0 12px rgba(0,229,255,0.3); }
  .btn-sm.danger { border-color: var(--pink); color: var(--pink); }
  .btn-sm.danger:hover { background: var(--pink); color: #fff; box-shadow: 0 0 12px rgba(255,64,129,0.3); }

  /* 搜索栏 */
  .search-bar {
    display: flex; gap: 8px; align-items: center; margin-bottom: 10px;
  }
  .search-bar input {
    flex: 1; min-width: 200px; max-width: 400px;
    padding: 8px 14px; border: 1px solid var(--border); border-radius: var(--radius);
    font-size: 13px; background: var(--card2); color: var(--text);
    transition: all 0.2s;
  }
  .search-bar input:focus { outline: none; border-color: var(--cyan); box-shadow: 0 0 10px rgba(0,229,255,0.15); }
  .search-bar .badge { font-size: 12px; color: var(--text2); }

  /* 统计条 */
  .stats { display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 10px; }
  .stat {
    background: var(--card); border: 1px solid var(--border);
    border-radius: var(--radius); padding: 8px 14px;
    font-size: 12px; box-shadow: var(--glow-cyan);
  }
  .stat b { color: var(--cyan); font-weight: 700; }

  .sel-info { font-size: 12px; color: var(--text2); margin-bottom: 6px; min-height: 18px; }
  .sel-info span { color: var(--pink); font-weight: 600; }

  /* 表格 */
  .table-wrap {
    background: var(--card); border: 1px solid var(--border);
    border-radius: 12px; box-shadow: var(--glow-cyan);
    overflow: auto; max-height: 46vh; margin-bottom: 10px;
  }
  table { width: 100%; border-collapse: collapse; font-size: 12px; white-space: nowrap; }
  thead { position: sticky; top: 0; z-index: 3; }
  th {
    background: linear-gradient(180deg, #1a2740, #111827);
    color: var(--cyan); padding: 9px 13px; text-align: left;
    cursor: pointer; user-select: none; font-weight: 600;
    border-bottom: 2px solid var(--cyan);
    font-family: "JetBrains Mono","Fira Code","Cascadia Code",monospace;
    transition: all 0.2s;
  }
  th:hover { background: #1e3355; color: #fff; }
  th.selected { background: var(--pink) !important; color: #fff !important; }
  th .col-type {
    display: block; font-size: 9px; font-weight: 400; color: var(--text2);
    margin-top: 1px; font-family: "Inter","Microsoft YaHei",sans-serif;
  }
  th.selected .col-type { color: rgba(255,255,255,0.7); }
  td {
    padding: 7px 13px; border-bottom: 1px solid rgba(30,45,69,0.6);
    cursor: pointer; transition: all 0.15s;
    font-family: "JetBrains Mono","Fira Code","Cascadia Code",monospace;
  }
  td:hover { background: rgba(0,229,255,0.04); }
  td.selected {
    background: rgba(255,64,129,0.12) !important;
    outline: 2px solid var(--pink); outline-offset: -2px;
    box-shadow: 0 0 12px rgba(255,64,129,0.2);
  }
  td.editing { padding: 0 !important; background: transparent !important; outline: 2px solid var(--cyan) !important; box-shadow: 0 0 18px rgba(0,229,255,0.3); }
  td input.cell-input {
    width: 100%; min-width: 100px; padding: 7px 13px;
    background: #0d1525; color: var(--cyan); border: none;
    font-family: inherit; font-size: inherit; outline: none;
  }
  /* 条件格式颜色 */
  td.cond-heat-0 { background: rgba(0, 123, 255, 0.18) !important; }
  td.cond-heat-1 { background: rgba(0, 150, 199, 0.18) !important; }
  td.cond-heat-2 { background: rgba(0, 176, 143, 0.18) !important; }
  td.cond-heat-3 { background: rgba(0, 203, 87, 0.18) !important; }
  td.cond-heat-4 { background: rgba(0, 230, 31, 0.18) !important; }

  /* 行选择 checkbox */
  .row-cb { width: 16px; height: 16px; accent-color: var(--cyan); cursor: pointer; }

  /* 面板 */
  .panel {
    background: var(--card); border: 1px solid var(--border);
    border-radius: 12px; box-shadow: var(--glow-cyan);
    padding: 14px 18px; margin-bottom: 10px;
  }
  .panel h3 {
    font-size: 14px; margin-bottom: 10px; color: var(--cyan);
    font-weight: 600; letter-spacing: 0.5px; cursor: pointer;
  }
  .panel h3 .toggle-icon { transition: transform 0.3s; display: inline-block; }
  .panel.collapsed .panel-body { display: none; }
  .panel.collapsed h3 .toggle-icon { transform: rotate(-90deg); }
  .panel h3 small { color: var(--text2); }
  .panel-body { transition: all 0.3s; }
  .op-grid { display: flex; gap: 10px; flex-wrap: wrap; align-items: flex-end; }
  .op-group { display: flex; flex-direction: column; gap: 4px; }
  .op-group label { font-size: 10px; color: var(--text2); font-weight: 500; text-transform: uppercase; letter-spacing: 0.5px; }
  .op-group input, .op-group select {
    padding: 7px 11px; border: 1px solid var(--border); border-radius: 6px;
    font-size: 12px; min-width: 120px; background: var(--card2); color: var(--text);
    transition: all 0.2s;
  }
  .op-group input:focus, .op-group select:focus {
    outline: none; border-color: var(--cyan); box-shadow: 0 0 8px rgba(0,229,255,0.15);
  }
  .op-group select option { background: var(--card2); color: var(--text); }

  /* 单元格操作 3 列网格 */
  .cell-ops { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 5px 12px; }
  .cell-ops .op-row {
    display: flex; gap: 5px; align-items: center; padding: 6px 10px;
    border-radius: 6px; background: var(--card2); min-height: 36px;
    border: 1px solid transparent; transition: border-color 0.2s;
  }
  .cell-ops .op-row:hover { border-color: rgba(0,229,255,0.2); }
  .cell-ops .op-row .op-label {
    font-size: 10px; color: var(--text2); min-width: 55px;
    font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;
  }
  .cell-ops .op-row input {
    width: 75px; padding: 5px 8px; border: 1px solid var(--border);
    border-radius: 5px; font-size: 11px; background: #0d1525; color: var(--text);
    transition: border-color 0.2s;
  }
  .cell-ops .op-row input:focus { border-color: var(--cyan); outline: none; }
  .cell-ops .op-row select {
    padding: 5px 8px; border: 1px solid var(--border);
    border-radius: 5px; font-size: 11px; background: #0d1525; color: var(--text);
  }

  /* 图表区 */
  .chart-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 14px; }
  .chart-box {
    background: var(--card2); border: 1px solid var(--border);
    border-radius: var(--radius); padding: 12px; position: relative;
    min-height: 280px;
  }
  .chart-box h4 { font-size: 12px; color: var(--text2); margin-bottom: 8px; }
  .chart-box canvas { max-height: 260px; }
  @media (max-width: 900px) { .chart-grid { grid-template-columns: 1fr; } }

  /* 列统计面板 */
  .col-stats-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(140px, 1fr)); gap: 8px; }
  .col-stat-item {
    background: var(--card2); padding: 10px; border-radius: var(--radius);
    border: 1px solid var(--border); text-align: center;
  }
  .col-stat-item .val { font-size: 20px; font-weight: 700; color: var(--cyan); }
  .col-stat-item .lbl { font-size: 10px; color: var(--text2); margin-top: 2px; text-transform: uppercase; }

  /* 模态 */
  .modal-overlay {
    position: fixed; inset: 0; background: rgba(0,0,0,0.7); z-index: 998;
    display: flex; align-items: center; justify-content: center;
  }
  .modal {
    background: var(--card); border: 1px solid var(--border);
    border-radius: 12px; box-shadow: 0 8px 48px rgba(0,0,0,0.6);
    padding: 24px; max-width: 800px; width: 90%; max-height: 80vh;
    overflow-y: auto; position: relative; z-index: 999;
  }
  .modal h3 { color: var(--cyan); margin-bottom: 14px; }
  .modal-close {
    position: absolute; top: 12px; right: 16px;
    cursor: pointer; font-size: 20px; color: var(--text2);
    background: none; border: none;
  }
  .modal-close:hover { color: var(--pink); }

  /* toast */
  .toast {
    position: fixed; top: 24px; right: 24px;
    background: linear-gradient(135deg, #1a2740, #0d1525);
    color: var(--cyan); border: 1px solid var(--cyan);
    padding: 12px 24px; border-radius: var(--radius);
    font-size: 13px; font-weight: 600;
    opacity: 0; transform: translateY(-20px);
    transition: all 0.3s cubic-bezier(0.4,0,0.2,1); z-index: 9999;
    box-shadow: 0 0 24px rgba(0,229,255,0.2);
  }
  .toast.show { opacity: 1; transform: translateY(0); }
  .toast.ok { border-color: var(--green); color: var(--green); box-shadow: 0 0 24px rgba(0,230,118,0.2); }
  .toast.err { border-color: var(--red); color: var(--red); box-shadow: 0 0 24px rgba(255,23,68,0.2); }

  .ctx-menu {
    position: fixed; background: var(--card); border: 1px solid var(--border);
    border-radius: var(--radius); box-shadow: 0 8px 32px rgba(0,0,0,0.5), var(--glow-cyan);
    z-index: 999; min-width: 200px; display: none; overflow: hidden;
  }
  .ctx-menu .item {
    padding: 10px 16px; cursor: pointer; font-size: 12px;
    border-bottom: 1px solid var(--border); transition: all 0.15s;
  }
  .ctx-menu .item:hover { background: var(--cyan); color: #000; }

  ::-webkit-scrollbar { width: 6px; height: 6px; }
  ::-webkit-scrollbar-track { background: var(--card); }
  ::-webkit-scrollbar-thumb { background: var(--border); border-radius: 3px; }
  ::-webkit-scrollbar-thumb:hover { background: var(--cyan); }

  /* 快捷键提示 */
  .kbd {
    display: inline-block; padding: 2px 6px; font-size: 10px;
    background: var(--card2); border: 1px solid var(--border);
    border-radius: 3px; font-family: "JetBrains Mono",monospace;
    color: var(--text2);
  }
</style>
</head>
<body>
<div class="app">
  <header>
    <h1>ACM 队伍数据管理平台</h1>
    <button class="btn btn-outline" onclick="undo()" title="Ctrl+Z">↩ <span class="kbd">Ctrl+Z</span></button>
    <button class="btn btn-outline" onclick="redo()" title="Ctrl+Y">↪ <span class="kbd">Ctrl+Y</span></button>
    <span style="color:#2a3a55;">|</span>
    <label class="btn btn-outline" style="cursor:pointer;">
      📂 导入 <input type="file" id="csvFile" accept=".csv,.xlsx" style="display:none">
    </label>
    <label class="btn btn-outline" style="cursor:pointer;">
      🔗 合并 <input type="file" id="mergeFile" accept=".csv,.xlsx" style="display:none">
    </label>
    <button class="btn btn-success" onclick="saveCSV()">💾 CSV</button>
    <button class="btn btn-success" onclick="saveExcel()">📊 Excel</button>
    <button class="btn btn-outline" onclick="saveReport()">📄 报告</button>
    <button class="btn btn-primary" onclick="resetDefault()">🔄 默认</button>
    <button class="btn btn-outline" onclick="clearSelection()">✕</button>
    <button class="btn btn-outline" onclick="openCharts()">📈</button>
    <button class="btn btn-outline" onclick="openInsights()">🧠 洞察</button>
    <button class="btn btn-outline" onclick="openBackups()">💿 备份</button>
    <button class="btn btn-outline" onclick="window.open('/print')">🖨 打印</button>
  </header>

  <!-- NL 查询 + 全局搜索 -->
  <div class="search-bar">
    <input id="globalSearch" type="text" placeholder="🔍 全局搜索..."
           oninput="onGlobalSearch()" onkeydown="if(event.key==='Escape'){this.value='';onGlobalSearch();}">
    <input id="nlQuery" type="text" placeholder="🤖 自然语言: 张三 AC数>100 前5行 按AC数降序"
           style="flex:2;max-width:500px;" onkeydown="if(event.key==='Enter')nlQuery()">
    <button class="btn-sm" onclick="nlQuery()">执行</button>
    <span class="badge" id="searchBadge"></span>
    <button class="btn-sm" onclick="clearSearch()">清除</button>
    <button class="btn-sm" onclick="openStats()">📊</button>
  </div>

  <div class="stats" id="stats"></div>
  <div class="sel-info">📌 当前选择: <span id="selText">无</span> &nbsp;| &nbsp;💡 双击单元格编辑 (Enter/Tab 确认, Esc 取消) &nbsp;| &nbsp;☑ 勾选行进行批量操作</div>

  <!-- 行操作工具栏 -->
  <div style="display:flex; gap:8px; margin-bottom:8px; flex-wrap:wrap; align-items:center;">
    <button class="btn btn-outline btn-sm" onclick="addRow()">➕ 新增行</button>
    <button class="btn btn-outline btn-sm danger" onclick="deleteSelectedRows()">🗑 删除选中行</button>
    <button class="btn btn-outline btn-sm" onclick="duplicateSelectedRows()">📋 复制选中行</button>
    <button class="btn btn-outline btn-sm" onclick="toggleAllRows()">☑ 全选/取消</button>
    <span style="font-size:11px;color:var(--text2);" id="rowSelBadge"></span>
    <span style="margin-left:auto;font-size:11px;color:var(--text2);">
      💡 <span class="kbd">Ctrl+S</span> 保存 &nbsp;
      <span class="kbd">Ctrl+Z</span> 撤销 &nbsp;
      <span class="kbd">Ctrl+Y</span> 重做 &nbsp;
      <span class="kbd">Ctrl+F</span> 搜索
    </span>
  </div>

  <div class="table-wrap">
    <table id="dataTable"><thead></thead><tbody></tbody></table>
  </div>

  <!-- 列级操作 -->
  <div class="panel">
    <h3 onclick="togglePanel(this)"><span class="toggle-icon">▼</span> 📊 列级操作 <small>（点击表头选中列）</small></h3>
    <div class="panel-body">
    <div class="op-grid">
      <button class="btn btn-danger btn-sm" onclick="colAction('delete')">🗑 删除列</button>
      <button class="btn btn-outline btn-sm" onclick="colAction('sort_asc')">⬆ 升序</button>
      <button class="btn btn-outline btn-sm" onclick="colAction('sort_desc')">⬇ 降序</button>
      <button class="btn btn-outline btn-sm" onclick="colAction('dedup')">🔍 去重</button>
      <button class="btn btn-outline btn-sm" onclick="colAction('count_values')">📈 值计数</button>
      <button class="btn btn-outline btn-sm" onclick="colAction('fill_empty')">🧹 填充空值</button>
      <button class="btn btn-outline btn-sm" onclick="colAction('clone')">📋 克隆列</button>
      <button class="btn btn-outline btn-sm" onclick="openColumnStats()">📊 列统计</button>
      <div class="op-group"><input id="col_rename_input" placeholder="新列名"></div>
      <button class="btn btn-outline btn-sm" onclick="colAction('rename')">✏ 重命名</button>
      <span style="margin:0 4px;color:#2a3a55;">|</span>
      <button class="btn btn-outline btn-sm" onclick="colAction('mask_col', {mode:'partial'})">🛡 脱敏(部分)</button>
      <button class="btn btn-outline btn-sm" onclick="colAction('mask_col', {mode:'stars'})">🛡 脱敏(全遮)</button>
      <button class="btn btn-outline btn-sm" onclick="colAction('mask_col', {mode:'hash'})">🛡 脱敏(哈希)</button>
      <button class="btn btn-outline btn-sm" onclick="colAction('reorder_alpha')">🔤 按字母重排</button>
    </div>
    </div>
  </div>

  <!-- 单元格操作 -->
  <div class="panel">
    <h3 onclick="togglePanel(this)"><span class="toggle-icon">▼</span> ✂ 单元格批量操作 <small>（点击单元格选中，Ctrl+点击多选）</small></h3>
    <div class="panel-body">
    <div class="cell-ops">
      <div class="op-row">
        <span class="op-label">删除字符</span>
        <input id="op_del_chars" placeholder="要删除的字符">
        <button class="btn-sm" onclick="cellAction('delete_chars')">执行</button>
      </div>
      <div class="op-row">
        <span class="op-label">大小写</span>
        <select id="op_case">
          <option value="upper">全部大写</option><option value="lower">全部小写</option>
          <option value="title">首字母大写</option><option value="capitalize">句首大写</option>
        </select>
        <button class="btn-sm" onclick="cellAction('case')">执行</button>
      </div>
      <div class="op-row">
        <span class="op-label">去除空格</span>
        <select id="op_strip">
          <option value="both">两端</option><option value="left">左侧</option>
          <option value="right">右侧</option><option value="all">全部空格</option>
        </select>
        <button class="btn-sm" onclick="cellAction('strip')">执行</button>
      </div>
      <div class="op-row">
        <span class="op-label">截取子串</span>
        <input id="op_slice_start" placeholder="start" style="width:54px">
        <input id="op_slice_end" placeholder="end" style="width:54px">
        <button class="btn-sm" onclick="cellAction('slice')">截取</button>
      </div>
      <div class="op-row">
        <span class="op-label">提取</span>
        <button class="btn-sm" onclick="cellAction('extract_digits')">数字</button>
        <button class="btn-sm" onclick="cellAction('extract_letters')">字母</button>
        <button class="btn-sm" onclick="cellAction('extract_chinese')">中文</button>
      </div>
      <div class="op-row">
        <span class="op-label">替换</span>
        <input id="op_replace_old" placeholder="查找">
        <input id="op_replace_new" placeholder="替换为">
        <button class="btn-sm" onclick="cellAction('replace')">执行</button>
      </div>
      <div class="op-row">
        <span class="op-label">添加</span>
        <input id="op_affix" placeholder="前缀/后缀">
        <button class="btn-sm" onclick="cellAction('prefix')">前</button>
        <button class="btn-sm" onclick="cellAction('suffix')">后</button>
      </div>
      <div class="op-row">
        <span class="op-label">分隔取段</span>
        <input id="op_split_sep" placeholder="分隔符">
        <input id="op_split_idx" placeholder="段号" style="width:52px">
        <button class="btn-sm" onclick="cellAction('split_take')">执行</button>
      </div>
      <div class="op-row">
        <span class="op-label">填充</span>
        <input id="op_pad_width" placeholder="长度" style="width:46px">
        <input id="op_pad_char" placeholder="字符" style="width:46px">
        <button class="btn-sm" onclick="cellAction('pad_left')">左</button>
        <button class="btn-sm" onclick="cellAction('pad_right')">右</button>
      </div>
      <div class="op-row">
        <span class="op-label">去除</span>
        <button class="btn-sm" onclick="cellAction('remove_digits')">数字</button>
        <button class="btn-sm" onclick="cellAction('remove_letters')">字母</button>
        <button class="btn-sm" onclick="cellAction('remove_chinese')">中文</button>
        <button class="btn-sm" onclick="cellAction('remove_punct')">标点</button>
      </div>
      <div class="op-row">
        <span class="op-label">正则替换</span>
        <input id="op_regex_pat" placeholder="正则">
        <input id="op_regex_rep" placeholder="替换为">
        <button class="btn-sm" onclick="cellAction('regex')">执行</button>
      </div>
      <div class="op-row">
        <span class="op-label">反转</span>
        <button class="btn-sm" onclick="cellAction('reverse')" style="flex:1">反转字符串</button>
      </div>
      <div class="op-row">
        <span class="op-label">转换</span>
        <button class="btn-sm" onclick="cellAction('to_number')">→数字</button>
        <button class="btn-sm" onclick="cellAction('to_text')">→文本</button>
        <button class="btn-sm danger" onclick="cellAction('clear')">清空</button>
      </div>
    </div>
    </div>
  </div>

  <!-- 筛选与计算列 -->
  <div class="panel">
    <h3 onclick="togglePanel(this)"><span class="toggle-icon">▼</span> 🔎 筛选 &amp; 计算列 &amp; 透视</h3>
    <div class="panel-body">
    <div class="op-grid">
      <div class="op-group"><label>条件列</label><select id="filter_col"></select></div>
      <div class="op-group"><label>操作</label>
        <select id="filter_op">
          <option value="contains">包含</option><option value="equals">等于</option>
          <option value="startswith">以…开头</option><option value="endswith">以…结尾</option>
          <option value="regex">正则匹配</option><option value="gt">&gt; (大于)</option><option value="lt">&lt; (小于)</option>
          <option value="not_empty">非空</option><option value="is_empty">为空</option>
        </select>
      </div>
      <div class="op-group"><label>值</label><input id="filter_val" placeholder="条件值"></div>
      <button class="btn btn-outline btn-sm" onclick="applyFilter()">🔍 筛选</button>
      <button class="btn btn-outline btn-sm" onclick="clearFilter()">✕ 恢复全部</button>
      <span style="margin:0 6px;color:#2a3a55;">|</span>
      <div class="op-group"><label>新列名</label><input id="calc_col_name" placeholder="列名"></div>
      <div class="op-group"><label>计算方式</label>
        <select id="calc_op">
          <option value="len">字符长度</option><option value="word_count">单词数</option>
          <option value="is_palindrome">是否回文</option><option value="digit_sum">各位数字和</option>
          <option value="upper">转大写</option><option value="lower">转小写</option>
          <option value="reverse">反转</option><option value="strip">去空格</option>
        </select>
      </div>
      <div class="op-group"><label>基于列</label><select id="calc_src_col"></select></div>
      <button class="btn btn-success btn-sm" onclick="addCalcCol()">➕ 添加</button>
      <span style="margin:0 6px;color:#2a3a55;">|</span>
      <div class="op-group"><label>公式列名</label><input id="formula_col_name" placeholder="列名"></div>
      <div class="op-group"><label>基于列</label><select id="formula_src_col"></select></div>
      <div class="op-group"><label>公式 (x=值, idx=行号)</label><input id="formula_expr" placeholder="例如: x*2+10"></div>
      <button class="btn btn-warn btn-sm" onclick="addFormulaCol()">🧮 公式</button>
      <span style="margin:0 6px;color:#2a3a55;">|</span>
      <div class="op-group"><label>透视行</label><select id="pivot_row_col"></select></div>
      <div class="op-group"><label>透视值</label><select id="pivot_val_col"></select></div>
      <div class="op-group"><label>聚合</label>
        <select id="pivot_agg">
          <option value="count">计数</option><option value="sum">求和</option>
          <option value="mean">平均</option><option value="max">最大</option><option value="min">最小</option>
        </select>
      </div>
      <button class="btn btn-warn btn-sm" onclick="applyPivot()">📋 透视表</button>
    </div>
    </div>
  </div>
</div>

<!-- 列统计模态 -->
<div class="modal-overlay" id="statsModal" style="display:none;">
  <div class="modal">
    <button class="modal-close" onclick="closeStats()">✕</button>
    <h3>📊 列统计 — <span id="statsColName" style="color:var(--pink);"></span></h3>
    <div class="col-stats-grid" id="colStatsGrid"></div>
    <canvas id="colStatsChart" style="max-height:280px; margin-top:16px;"></canvas>
  </div>
</div>

<!-- 可视化模态 -->
<div class="modal-overlay" id="chartsModal" style="display:none;">
  <div class="modal" style="max-width:1000px;">
    <button class="modal-close" onclick="closeCharts()">✕</button>
    <h3>📈 数据可视化</h3>
    <div class="op-grid" style="margin-bottom:12px;">
      <div class="op-group"><label>X轴列</label><select id="chartXCol"></select></div>
      <div class="op-group"><label>Y轴列</label><select id="chartYCol"></select></div>
      <div class="op-group"><label>图表类型</label>
        <select id="chartType">
          <option value="bar">柱状图</option><option value="line">折线图</option>
          <option value="pie">饼图</option><option value="radar">雷达图</option>
          <option value="scatter">散点图</option>
        </select>
      </div>
      <button class="btn btn-primary btn-sm" onclick="renderChart()">🔄 刷新图表</button>
    </div>
    <div class="chart-grid">
      <div class="chart-box"><h4>主图表</h4><canvas id="mainChart"></canvas></div>
      <div class="chart-box"><h4>汇总图表</h4><canvas id="summaryChart"></canvas></div>
    </div>
  </div>
</div>

<!-- 洞察模态 -->
<div class="modal-overlay" id="insightsModal" style="display:none;">
  <div class="modal" style="max-width:900px;">
    <button class="modal-close" onclick="closeInsights()">✕</button>
    <h3>🧠 数据洞察报告</h3>
    <div id="insightsContent"></div>
  </div>
</div>

<!-- 备份模态 -->
<div class="modal-overlay" id="backupsModal" style="display:none;">
  <div class="modal" style="max-width:900px;">
    <button class="modal-close" onclick="closeBackups()">✕</button>
    <h3>💿 数据备份管理</h3>
    <div style="margin-bottom:12px;">
      <button class="btn btn-primary btn-sm" onclick="createBackup()">➕ 创建备份</button>
      <button class="btn btn-outline btn-sm" onclick="saveToDB()">💾 存SQLite</button>
      <button class="btn btn-outline btn-sm" onclick="loadFromDB()">📥 读SQLite</button>
      <label class="btn btn-outline btn-sm" style="cursor:pointer;">
        📤 上传对比 <input type="file" id="diffFile" accept=".csv,.xlsx" style="display:none;">
      </label>
    </div>
    <div id="diffSection" style="display:none;"><div id="diffContent"></div></div>
    <div id="backupList"></div>
  </div>
</div>

<div class="toast" id="toast"></div>
<div class="ctx-menu" id="ctxMenu"></div>

<script>
// ===================== 状态 =====================
let selectedCol = null;
let selectedCells = new Set();
let selectedRows = new Set();  // 行复选框选中的行索引
let _cols = [], _rows = [], _types = {}, _condFormat = {};
let _editingCell = null;
let _clickTimer = null;
let _chartMain = null, _chartSummary = null, _colStatsChart = null;
let _searchTerm = "";

// ===================== 渲染表格 =====================
async function loadTable() {
  const url = _searchTerm ? `/api/data?search=${encodeURIComponent(_searchTerm)}` : "/api/data";
  const res = await fetch(url);
  const j = await res.json();
  _cols = j.columns; _rows = j.rows; _types = j.types; _condFormat = j.cond_format || {};
  _canUndo = j.can_undo; _canRedo = j.can_redo;

  document.getElementById("stats").innerHTML =
    `<div class="stat">行数 <b>${_rows.length}</b></div>
     <div class="stat">列数 <b>${_cols.length}</b></div>` +
    _cols.map(c => `<div class="stat">${c} <b>${_types[c]}</b></div>`).join("");

  // 表头
  let th = "<tr><th>☑</th><th>#</th>";
  for (const c of _cols) {
    th += `<th class="${selectedCol===c?'selected':''}" draggable="true" ondragstart="dragCol(event,'${escAttr(c)}')" ondragover="dragOver(event)" ondrop="dropCol(event,'${escAttr(c)}')" onclick="selectCol('${escAttr(c)}')">${escHtml(c)}<span class="col-type">${_types[c]}</span></th>`;
  }
  th += "</tr>";
  document.querySelector("#dataTable thead").innerHTML = th;

  // 表体
  let tb = "";
  for (let i = 0; i < _rows.length; i++) {
    const rowChecked = selectedRows.has(i) ? " checked" : "";
    tb += `<tr>`;
    tb += `<td style="padding:4px 8px;text-align:center;"><input class="row-cb" type="checkbox" ${rowChecked} onchange="toggleRow(${i}, this.checked)" onclick="event.stopPropagation();"></td>`;
    tb += `<td style="color:var(--text2);cursor:default;font-size:10px;padding:4px 8px;">${i}</td>`;
    for (const c of _cols) {
      const key = i + "," + c;
      const sel = selectedCells.has(key) ? " selected" : "";
      const edit = _editingCell && _editingCell.row === i && _editingCell.col === c ? " editing" : "";
      const raw = _rows[i][c] ?? "";
      // 条件格式
      let condClass = "";
      if (_condFormat[c] && _condFormat[c][i] !== undefined) {
        condClass = " cond-heat-" + _condFormat[c][i];
      }
      if (_editingCell && _editingCell.row === i && _editingCell.col === c) {
        tb += `<td class="${sel}${edit}${condClass}"><input class="cell-input" id="cellInput" value="${escAttr(String(raw))}"></td>`;
      } else {
        tb += `<td class="${sel}${edit}${condClass}" onclick="selectCell(event,${i},'${escAttr(c)}')" ondblclick="startEdit(event,${i},'${escAttr(c)}')">${escHtml(String(raw))}</td>`;
      }
    }
    tb += "</tr>";
  }
  document.querySelector("#dataTable tbody").innerHTML = tb;

  if (_editingCell) {
    const inp = document.getElementById("cellInput");
    if (inp) {
      inp.focus(); inp.select();
      inp.addEventListener("keydown", onEditKey);
      inp.addEventListener("blur", commitEdit);
    }
  }

  updateRowBadge();
  updateSearchBadge();

  // 填充下拉
  const opts = _cols.map(c => `<option value="${escAttr(c)}">${escHtml(c)}</option>`).join("");
  document.getElementById("filter_col").innerHTML = opts;
  document.getElementById("calc_src_col").innerHTML = opts;
  document.getElementById("pivot_row_col").innerHTML = opts;
  document.getElementById("pivot_val_col").innerHTML = opts;
  document.getElementById("chartXCol").innerHTML = opts;
  document.getElementById("chartYCol").innerHTML = opts;
  document.getElementById("formula_src_col").innerHTML = opts;
}

function escHtml(s) { const d=document.createElement("div"); d.textContent = String(s); return d.innerHTML; }
function escAttr(s) { return String(s).replace(/&/g,"&amp;").replace(/"/g,"&quot;").replace(/'/g,"&#39;").replace(/</g,"&lt;").replace(/>/g,"&gt;"); }

// ===================== 行选择 =====================
function toggleRow(idx, checked) {
  checked ? selectedRows.add(idx) : selectedRows.delete(idx);
  updateRowBadge(); autoSave();
}

function toggleAllRows() {
  if (selectedRows.size === _rows.length) {
    selectedRows.clear();
  } else {
    for (let i = 0; i < _rows.length; i++) selectedRows.add(i);
  }
  loadTable();
}

function updateRowBadge() {
  const n = selectedRows.size;
  document.getElementById("rowSelBadge").textContent = n > 0 ? `已选 ${n} 行` : "";
}

// ===================== 行操作 =====================
async function addRow() {
  const res = await fetch("/api/row/add", { method:"POST",
    headers:{"Content-Type":"application/json"},
    body: JSON.stringify({})
  });
  const j = await res.json();
  toast(j.message, j.ok?"ok":"err");
  loadTable();
}

async function deleteSelectedRows() {
  if (selectedRows.size === 0) return toast("请先勾选要删除的行", "err");
  if (!confirm(`确定删除 ${selectedRows.size} 行？此操作可撤销。`)) return;
  const res = await fetch("/api/row/delete", { method:"POST",
    headers:{"Content-Type":"application/json"},
    body: JSON.stringify({ rows: [...selectedRows] })
  });
  const j = await res.json();
  selectedRows.clear();
  toast(j.message, j.ok?"ok":"err");
  loadTable();
}

async function duplicateSelectedRows() {
  if (selectedRows.size === 0) return toast("请先勾选要复制的行", "err");
  const res = await fetch("/api/row/duplicate", { method:"POST",
    headers:{"Content-Type":"application/json"},
    body: JSON.stringify({ rows: [...selectedRows] })
  });
  const j = await res.json();
  selectedRows.clear();
  toast(j.message, j.ok?"ok":"err");
  loadTable();
}

// ===================== 撤销/重做 =====================
async function undo() {
  const res = await fetch("/api/undo", { method:"POST" });
  const j = await res.json();
  toast(j.message, j.ok?"ok":"err");
  loadTable();
}

async function redo() {
  const res = await fetch("/api/redo", { method:"POST" });
  const j = await res.json();
  toast(j.message, j.ok?"ok":"err");
  loadTable();
}

// ===================== 全局搜索 =====================
function onGlobalSearch() {
  _searchTerm = document.getElementById("globalSearch").value.trim();
  autoSave(); loadTable();
}

function clearSearch() {
  document.getElementById("globalSearch").value = "";
  _searchTerm = "";
  loadTable();
}

function updateSearchBadge() {
  const badge = document.getElementById("searchBadge");
  if (_searchTerm) {
    badge.textContent = `搜索 "${_searchTerm}" — ${_rows.length} 条结果`;
  } else {
    badge.textContent = "";
  }
}

// ===================== 内联编辑 =====================
function startEdit(e, row, col) {
  e.preventDefault();
  if (_clickTimer) { clearTimeout(_clickTimer); _clickTimer = null; }
  _editingCell = { row, col, origVal: _rows[row]?.[col] ?? "" };
  selectedCells.clear();
  selectedCol = null;
  loadTable();
}

function onEditKey(e) {
  if (e.key === "Enter") {
    e.preventDefault();
    const cell = _editingCell;
    commitEdit();
    if (cell) {
      const nextRow = Math.min(cell.row + 1, _rows.length - 1);
      startEdit(e, nextRow, cell.col);
    }
  } else if (e.key === "Tab") {
    e.preventDefault();
    const cell = _editingCell;
    commitEdit();
    if (cell) {
      const idx = _cols.indexOf(cell.col);
      const nextIdx = e.shiftKey ? idx - 1 : idx + 1;
      if (nextIdx >= 0 && nextIdx < _cols.length) {
        startEdit(e, cell.row, _cols[nextIdx]);
      }
    }
  } else if (e.key === "Escape") {
    e.preventDefault();
    _editingCell = null;
    loadTable();
  }
}

async function commitEdit() {
  if (!_editingCell) return;
  const inp = document.getElementById("cellInput");
  const newVal = inp ? inp.value : _editingCell.origVal;
  const { row, col, origVal } = _editingCell;
  _editingCell = null;
  if (newVal === origVal) { loadTable(); return; }
  const res = await fetch("/edit_cell", {
    method: "POST",
    headers: {"Content-Type":"application/json"},
    body: JSON.stringify({ row, col, value: newVal })
  });
  const j = await res.json();
  if (j.ok) {
    toast("已保存", "ok");
    if (_rows[row]) _rows[row][col] = newVal;
  } else {
    toast(j.message || "保存失败", "err");
  }
  loadTable();
}

// ===================== 选择逻辑 =====================
function selectCol(col) {
  _editingCell = null;
  selectedCol = (selectedCol === col) ? null : col;
  document.getElementById("selText").innerHTML = selectedCol ? `列 <b>${escHtml(selectedCol)}</b>` : "无";
  autoSave(); loadTable();
}

function selectCell(e, rowIdx, col) {
  if (_editingCell) { commitEdit(); return; }
  const key = rowIdx + "," + col;
  if (e.ctrlKey || e.metaKey) {
    if (_clickTimer) { clearTimeout(_clickTimer); _clickTimer = null; }
    selectedCells.has(key) ? selectedCells.delete(key) : selectedCells.add(key);
  } else {
    if (_clickTimer) { clearTimeout(_clickTimer); _clickTimer = null; return; }
    _clickTimer = setTimeout(() => {
      _clickTimer = null;
      selectedCells.clear(); selectedCells.add(key); selectedCol = null;
      const arr = [...selectedCells];
      const preview = arr.length <= 3 ? arr.join(", ") : arr.slice(0,3).join(", ") + ` …等${arr.length}个`;
      document.getElementById("selText").innerHTML = `单元格: ${preview}`;
      loadTable();
    }, 280);
    return;
  }
  selectedCol = null;
  const arr = [...selectedCells];
  const preview = arr.length <= 3 ? arr.join(", ") : arr.slice(0,3).join(", ") + ` …等${arr.length}个`;
  document.getElementById("selText").innerHTML = `单元格: ${preview}`;
  autoSave(); loadTable();
}

function clearSelection() {
  _editingCell = null;
  selectedCol = null; selectedCells.clear(); selectedRows.clear();
  document.getElementById("selText").innerHTML = "无";
  autoSave(); loadTable();
}

// ===================== 文件上传 (CSV/Excel) =====================
document.getElementById("csvFile").addEventListener("change", async (e) => {
  const file = e.target.files[0];
  if (!file) return;
  const fd = new FormData(); fd.append("file", file);
  const res = await fetch("/upload", { method:"POST", body:fd });
  const j = await res.json();
  toast(j.message || "加载完成", j.ok?"ok":"err");
  clearSelection(); loadTable();
  e.target.value = "";
});

// ===================== 合并数据集 =====================
document.getElementById("mergeFile").addEventListener("change", async (e) => {
  const file = e.target.files[0];
  if (!file) return;
  const mode = prompt("合并模式:\n- concat (纵向追加)\n- join (横向关联, 需同名列)", "concat") || "concat";
  let joinCol = "";
  if (mode === "join") {
    joinCol = prompt("请输入两个数据集共有列名用于关联:", _cols[0] || "");
    if (!joinCol) { e.target.value = ""; return; }
  }
  const fd = new FormData(); fd.append("file", file);
  fd.append("mode", mode);
  if (joinCol) fd.append("join_col", joinCol);
  const res = await fetch("/api/merge", { method:"POST", body:fd });
  const j = await res.json();
  toast(j.message || "合并完成", j.ok?"ok":"err");
  clearSelection(); loadTable();
  e.target.value = "";
});

// ===================== 列操作 =====================
async function colAction(action, extraParams) {
  if (!selectedCol) return toast("请先点击表头选中一列", "err");
  const params = extraParams ? {...extraParams} : {};
  if (action === "rename") params.new_name = document.getElementById("col_rename_input").value.trim();
  if (action === "reorder_alpha") {
    const sorted = [..._cols].sort((a,b) => a.localeCompare(b, 'zh-CN'));
    params.order = sorted;
    action = "reorder_cols";
  }
  const res = await fetch("/process", {
    method:"POST",
    headers:{"Content-Type":"application/json"},
    body: JSON.stringify({ action, col: selectedCol, params })
  });
  const j = await res.json();
  toast(j.message || "完成", j.ok?"ok":"err");
  loadTable();
}

// ===================== 单元格批量操作 =====================
async function cellAction(action) {
  if (selectedCells.size === 0) return toast("请先选中单元格（Ctrl+点击可多选）", "err");
  const getVal = id => document.getElementById(id)?.value || "";
  const params = {};
  switch (action) {
    case "delete_chars": params.chars = getVal("op_del_chars"); break;
    case "replace": params.old = getVal("op_replace_old"); params.new = getVal("op_replace_new"); break;
    case "case": params.mode = getVal("op_case"); break;
    case "strip": params.mode = getVal("op_strip"); break;
    case "prefix": case "suffix": params.text = getVal("op_affix"); break;
    case "slice": params.start = getVal("op_slice_start"); params.end = getVal("op_slice_end"); break;
    case "regex": params.pattern = getVal("op_regex_pat"); params.repl = getVal("op_regex_rep"); break;
    case "split_take": params.sep = getVal("op_split_sep"); params.idx = getVal("op_split_idx"); break;
    case "pad_left": case "pad_right": params.width = getVal("op_pad_width"); params.char = getVal("op_pad_char"); break;
  }
  const cells = [...selectedCells].map(k => { const p = k.split(",",2); return [parseInt(p[0]), p[1]]; });
  const res = await fetch("/process", {
    method:"POST",
    headers:{"Content-Type":"application/json"},
    body: JSON.stringify({ action, cells, params })
  });
  const j = await res.json();
  toast(j.message || "完成", j.ok?"ok":"err");
  loadTable();
}

// ===================== 筛选 =====================
async function applyFilter() {
  const col = document.getElementById("filter_col").value;
  const op = document.getElementById("filter_op").value;
  const val = document.getElementById("filter_val").value;
  const res = await fetch("/process", {
    method:"POST",
    headers:{"Content-Type":"application/json"},
    body: JSON.stringify({ action:"filter", col, params:{op, val} })
  });
  const j = await res.json();
  toast(j.message || "筛选完成", j.ok?"ok":"err");
  loadTable();
}

async function clearFilter() {
  await fetch("/process", { method:"POST",
    headers:{"Content-Type":"application/json"},
    body: JSON.stringify({ action:"clear_filter" })
  });
  loadTable();
  toast("已恢复全部数据", "ok");
}

// ===================== 计算列 =====================
async function addCalcCol() {
  const name = document.getElementById("calc_col_name").value.trim();
  const op = document.getElementById("calc_op").value;
  const src = document.getElementById("calc_src_col").value;
  if (!name) return toast("请输入新列名", "err");
  const res = await fetch("/process", {
    method:"POST",
    headers:{"Content-Type":"application/json"},
    body: JSON.stringify({ action:"calc_col", col:src, params:{op, new_col:name} })
  });
  const j = await res.json();
  toast(j.message || "已添加", j.ok?"ok":"err");
  loadTable();
}

// ===================== 透视表 =====================
async function applyPivot() {
  const rowCol = document.getElementById("pivot_row_col").value;
  const valCol = document.getElementById("pivot_val_col").value;
  const agg = document.getElementById("pivot_agg").value;
  if (!rowCol || !valCol) return toast("请选择透视行和值列", "err");
  const res = await fetch("/process", {
    method:"POST",
    headers:{"Content-Type":"application/json"},
    body: JSON.stringify({ action:"pivot", col:rowCol, params:{val_col:valCol, agg} })
  });
  const j = await res.json();
  toast(j.message || "透视完成", j.ok?"ok":"err");
  loadTable();
}

// ===================== 列统计 =====================
async function openColumnStats() {
  if (!selectedCol) return toast("请先点击表头选中一列", "err");
  openStatsForCol(selectedCol);
}

async function openStats() {
  // 打开第一个数值列的统计
  const numCol = _cols.find(c => _types[c] === "数值") || _cols[0];
  if (numCol) openStatsForCol(numCol);
}

async function openStatsForCol(col) {
  document.getElementById("statsColName").textContent = col;
  const res = await fetch(`/api/stats/${encodeURIComponent(col)}`);
  const j = await res.json();
  if (!j.ok) return toast(j.message, "err");

  let html = "";
  for (const [k, v] of Object.entries(j.stats)) {
    html += `<div class="col-stat-item"><div class="val">${v}</div><div class="lbl">${k}</div></div>`;
  }
  document.getElementById("colStatsGrid").innerHTML = html;
  document.getElementById("statsModal").style.display = "flex";

  // 小图表
  if (_colStatsChart) _colStatsChart.destroy();
  const ctx = document.getElementById("colStatsChart").getContext("2d");
  if (j.histogram && j.histogram.labels && j.histogram.labels.length > 0) {
    _colStatsChart = new Chart(ctx, {
      type: "bar",
      data: {
        labels: j.histogram.labels,
        datasets: [{
          label: col + " 分布",
          data: j.histogram.values,
          backgroundColor: "rgba(0,229,255,0.3)",
          borderColor: "#00e5ff",
          borderWidth: 1
        }]
      },
      options: {
        responsive: true,
        plugins: { legend: { display: false } },
        scales: {
          y: { beginAtZero: true, grid: { color: "rgba(30,45,69,0.5)" }, ticks: { color: "#94a3b8" } },
          x: { grid: { color: "rgba(30,45,69,0.5)" }, ticks: { color: "#94a3b8" } }
        }
      }
    });
  }
}

function closeStats() {
  document.getElementById("statsModal").style.display = "none";
  if (_colStatsChart) { _colStatsChart.destroy(); _colStatsChart = null; }
}

// ===================== 可视化 =====================
function openCharts() {
  document.getElementById("chartsModal").style.display = "flex";
  setTimeout(() => renderChart(), 200);
}

function closeCharts() {
  document.getElementById("chartsModal").style.display = "none";
  if (_chartMain) { _chartMain.destroy(); _chartMain = null; }
  if (_chartSummary) { _chartSummary.destroy(); _chartSummary = null; }
}

async function renderChart() {
  const xCol = document.getElementById("chartXCol").value || _cols[0];
  const yCol = document.getElementById("chartYCol").value || (_cols.find(c => _types[c]==="数值") || _cols[0]);
  const chartType = document.getElementById("chartType").value;

  const labels = _rows.map(r => r[xCol] ?? "");
  const values = _rows.map(r => {
    const v = parseFloat(r[yCol]);
    return isNaN(v) ? 0 : v;
  });

  const colors = ["#00e5ff","#ff4081","#00e676","#ffab00","#7c4dff","#18ffff","#ff6e40","#69f0ae"];

  // 主图表
  if (_chartMain) _chartMain.destroy();
  const ctx1 = document.getElementById("mainChart").getContext("2d");
  _chartMain = new Chart(ctx1, {
    type: chartType === "scatter" ? "bar" : chartType,
    data: {
      labels,
      datasets: [{
        label: yCol,
        data: values,
        backgroundColor: labels.map((_,i) => colors[i % colors.length] + "44"),
        borderColor: colors[0],
        borderWidth: 1,
        tension: 0.3
      }]
    },
    options: {
      responsive: true,
      plugins: {
        legend: { labels: { color: "#94a3b8" } },
        title: { display: true, text: `${yCol} 按 ${xCol}`, color: "#e2e8f0" }
      },
      scales: chartType !== "pie" ? {
        y: { beginAtZero: true, grid: { color: "rgba(30,45,69,0.5)" }, ticks: { color: "#94a3b8" } },
        x: { grid: { color: "rgba(30,45,69,0.5)" }, ticks: { color: "#94a3b8" } }
      } : {}
    }
  });

  // 汇总图表 - 按队伍聚合
  if (_chartSummary) _chartSummary.destroy();
  const ctx2 = document.getElementById("summaryChart").getContext("2d");
  // 聚合: 按 xCol 分组 sum yCol
  const aggMap = {};
  for (let i = 0; i < labels.length; i++) {
    const k = labels[i];
    aggMap[k] = (aggMap[k] || 0) + values[i];
  }
  const aggLabels = Object.keys(aggMap);
  const aggValues = Object.values(aggMap);
  _chartSummary = new Chart(ctx2, {
    type: "bar",
    data: {
      labels: aggLabels,
      datasets: [{
        label: `${yCol} (聚合)`,
        data: aggValues,
        backgroundColor: aggLabels.map((_,i) => colors[i % colors.length] + "66"),
        borderColor: "#ff4081",
        borderWidth: 1
      }]
    },
    options: {
      responsive: true,
      plugins: {
        legend: { labels: { color: "#94a3b8" } },
        title: { display: true, text: `${yCol} 按 ${xCol} 聚合`, color: "#e2e8f0" }
      },
      scales: {
        y: { beginAtZero: true, grid: { color: "rgba(30,45,69,0.5)" }, ticks: { color: "#94a3b8" } },
        x: { grid: { color: "rgba(30,45,69,0.5)" }, ticks: { color: "#94a3b8" } }
      }
    }
  });
}

// ===================== NL 查询 =====================
async function nlQuery() {
  const q = document.getElementById("nlQuery").value.trim();
  if (!q) return toast("请输入查询语句", "err");
  const res = await fetch(`/api/nlquery?q=${encodeURIComponent(q)}`);
  const j = await res.json();
  if (!j.ok) return toast(j.message, "err");
  toast(`NL查询: ${j.explanation} — ${j.result_count} 条结果`, "ok");
  _cols = j.columns;
  _rows = j.rows;
  _types = {};
  _condFormat = {};
  for (const c of _cols) {
    try { parseFloat(_rows[0]?.[c]); _types[c] = "数值"; } catch(e) { _types[c] = "文本"; }
  }
  selectedCells.clear(); selectedRows.clear(); selectedCol = null;
  loadTable();
}

// ===================== 备份管理 =====================
async function openBackups() {
  document.getElementById("backupsModal").style.display = "flex";
  refreshBackupList();
}

async function refreshBackupList() {
  const res = await fetch("/api/backup/list");
  const j = await res.json();
  let html = "";
  if (j.backups.length === 0) {
    html = '<div style="color:var(--text2);padding:20px;text-align:center;">暂无备份</div>';
  } else {
    html = '<table><thead><tr><th>名称</th><th>大小</th><th>时间</th><th>操作</th></tr></thead><tbody>';
    for (const b of j.backups) {
      const safeName = escAttr(b.name);
      html += `<tr><td style="font-family:monospace;">${escHtml(b.name)}</td><td>${b.size}</td><td>${b.modified}</td>
        <td>
          <button class="btn-sm" onclick="restoreBackup('${safeName}')">恢复</button>
          <button class="btn-sm danger" onclick="deleteBackup('${safeName}')">删除</button>
          <button class="btn-sm" onclick="diffBackup('${safeName}')">对比</button>
        </td></tr>`;
    }
    html += "</tbody></table>";
  }
  document.getElementById("backupList").innerHTML = html;
}

async function createBackup() {
  const name = prompt("备份名称 (留空=自动生成时间戳):", "");
  const res = await fetch("/api/backup/create", {
    method:"POST",
    headers:{"Content-Type":"application/json"},
    body: JSON.stringify({ name: name || "" })
  });
  const j = await res.json();
  toast(j.message, j.ok?"ok":"err");
  refreshBackupList();
}

async function restoreBackup(name) {
  if (!confirm(`确定恢复备份「${name}」？当前数据将被替换。`)) return;
  const res = await fetch("/api/backup/restore", {
    method:"POST",
    headers:{"Content-Type":"application/json"},
    body: JSON.stringify({ name })
  });
  const j = await res.json();
  toast(j.message, j.ok?"ok":"err");
  if (j.ok) { clearSelection(); loadTable(); }
}

async function deleteBackup(name) {
  if (!confirm(`确定删除备份「${name}」？此操作不可撤销。`)) return;
  const res = await fetch("/api/backup/delete", {
    method:"POST",
    headers:{"Content-Type":"application/json"},
    body: JSON.stringify({ name })
  });
  const j = await res.json();
  toast(j.message, j.ok?"ok":"err");
  refreshBackupList();
}

async function diffBackup(name) {
  const res = await fetch(`/api/diff?backup=${encodeURIComponent(name)}`, { method:"POST" });
  const j = await res.json();
  if (!j.ok) return toast(j.message, "err");
  document.getElementById("diffContent").innerHTML =
    `<div style="background:var(--card2);padding:14px;border-radius:8px;margin:8px 0;">
      <div style="font-weight:600;color:var(--cyan);">对比结果: ${escHtml(j.source)}</div>
      <div style="margin-top:8px;display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;">
        <div class="col-stat-item"><div class="val" style="color:var(--green);">${j.only_in_current}</div><div class="lbl">仅在当前</div></div>
        <div class="col-stat-item"><div class="val" style="color:var(--amber);">${j.in_both}</div><div class="lbl">共同行</div></div>
        <div class="col-stat-item"><div class="val" style="color:var(--pink);">${j.only_in_source}</div><div class="lbl">仅在来源</div></div>
      </div>
      <div style="margin-top:4px;font-size:11px;color:var(--text2);">当前 ${j.current_rows}行 | 来源 ${j.source_rows}行 | ${j.common_cols}个共同列 | 差异值约${j.modified_values}个</div>
    </div>`;
  document.getElementById("diffSection").style.display = "block";
  document.getElementById("backupsModal").style.display = "flex";
}

function closeBackups() {
  document.getElementById("backupsModal").style.display = "none";
  document.getElementById("diffSection").style.display = "none";
}

async function saveToDB() {
  const res = await fetch("/api/db/save", { method:"POST" });
  const j = await res.json();
  toast(j.message, j.ok?"ok":"err");
}

async function loadFromDB() {
  if (!confirm("从数据库加载将替换当前数据，是否继续？")) return;
  const res = await fetch("/api/db/load", { method:"POST" });
  const j = await res.json();
  toast(j.message, j.ok?"ok":"err");
  if (j.ok) { clearSelection(); loadTable(); }
}

// Diff 文件上传 — 用 MutationObserver 确保 DOM 加载后绑定
(function bindDiffUpload() {
  const tryBind = () => {
    const diffInput = document.getElementById("diffFile");
    if (diffInput && !diffInput._bound) {
      diffInput._bound = true;
      diffInput.addEventListener("change", async (e) => {
        const file = e.target.files[0];
        if (!file) return;
        const fd = new FormData(); fd.append("file", file);
        const res = await fetch("/api/diff", { method:"POST", body: fd });
        const j = await res.json();
        if (!j.ok) return toast(j.message, "err");
        document.getElementById("diffContent").innerHTML =
          `<div style="background:var(--card2);padding:14px;border-radius:8px;margin:8px 0;">
            <div style="font-weight:600;color:var(--cyan);">对比: ${escHtml(j.source)}</div>
            <div style="margin-top:8px;display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;">
              <div class="col-stat-item"><div class="val" style="color:var(--green);">${j.only_in_current}</div><div class="lbl">仅在当前</div></div>
              <div class="col-stat-item"><div class="val" style="color:var(--amber);">${j.in_both}</div><div class="lbl">共同行</div></div>
              <div class="col-stat-item"><div class="val" style="color:var(--pink);">${j.only_in_source}</div><div class="lbl">仅在来源</div></div>
            </div>
            <div style="margin-top:4px;font-size:11px;color:var(--text2);">${j.summary}</div>
          </div>`;
        document.getElementById("diffSection").style.display = "block";
        document.getElementById("backupsModal").style.display = "flex";
        e.target.value = "";
      });
    }
  };
  if (document.readyState === "loading") {
    document.addEventListener("DOMContentLoaded", tryBind);
  } else {
    tryBind();
    // 也延迟重试
    setTimeout(tryBind, 300);
  }
})();

// ===================== 公式列 =====================
async function addFormulaCol() {
  const name = document.getElementById("formula_col_name").value.trim();
  const src = document.getElementById("formula_src_col").value;
  const formula = document.getElementById("formula_expr").value.trim();
  if (!name) return toast("请输入公式列名", "err");
  if (!formula) return toast("请输入公式", "err");
  const res = await fetch("/process", {
    method:"POST",
    headers:{"Content-Type":"application/json"},
    body: JSON.stringify({ action:"formula_col", col:src, params:{new_col:name, formula} })
  });
  const j = await res.json();
  toast(j.message || "已添加", j.ok?"ok":"err");
  loadTable();
}

// ===================== 数据洞察 =====================
async function openInsights() {
  document.getElementById("insightsModal").style.display = "flex";
  const res = await fetch("/api/insights");
  const j = await res.json();
  if (!j.ok) return toast(j.message, "err");
  const gradeColors = {"优秀":"#00e676","良好":"#ffab00","一般":"#ff9100","较差":"#ff1744"};
  let html = `<div style="display:flex;gap:16px;flex-wrap:wrap;margin-bottom:16px;align-items:center;">
    <div style="font-size:48px;font-weight:900;color:${gradeColors[j.quality_grade]||'#00e5ff'};">${j.quality_score}分</div>
    <div><div style="font-size:16px;color:var(--text);">数据质量: <b style="color:${gradeColors[j.quality_grade]};">${j.quality_grade}</b></div>
    <div style="font-size:12px;color:var(--text2);">${j.total_rows}行 × ${j.total_cols}列 | ${j.num_cols}个数值列 ${j.text_cols_count}个文本列</div></div></div>`;

  html += j.insights.map((ins, i) => {
    const icon = ins.type === "warning" ? "⚠️" : ins.type === "overview" ? "📋" : ins.type === "correlation" ? "🔗" : ins.type === "info" ? "ℹ️" : ins.type === "category" ? "🏷" : "💡";
    const borderColor = ins.type === "warning" ? "var(--amber)" : ins.type === "correlation" ? "var(--pink)" : "var(--cyan)";
    return `<div style="background:var(--card2);border-left:3px solid ${borderColor};padding:10px 14px;margin-bottom:8px;border-radius:4px;">
      <div style="font-weight:600;font-size:13px;">${icon} ${ins.title}</div>
      <div style="font-size:11px;color:var(--text2);margin-top:4px;">${ins.text}</div></div>`;
  }).join("");

  document.getElementById("insightsContent").innerHTML = html;
}

function closeInsights() {
  document.getElementById("insightsModal").style.display = "none";
}

// ===================== HTML 报告 =====================
function saveReport() { window.location.href = "/save_html_report"; }

// ===================== 保存 & 重置 =====================
function saveCSV() { window.location.href = "/save"; }
function saveExcel() { window.location.href = "/save_excel"; }

// ===================== 自动保存 (localStorage) =====================
let _autoSaveTimer = null;
function autoSave() {
  clearTimeout(_autoSaveTimer);
  _autoSaveTimer = setTimeout(async () => {
    try {
      const state = {
        selectedCells: [...selectedCells],
        selectedRows: [...selectedRows],
        selectedCol,
        searchTerm: _searchTerm,
        timestamp: Date.now()
      };
      localStorage.setItem("acm_data_state", JSON.stringify(state));
    } catch(e) {}
  }, 1000);
}

function loadAutoSave() {
  try {
    const raw = localStorage.getItem("acm_data_state");
    if (!raw) return;
    const state = JSON.parse(raw);
    if (Date.now() - state.timestamp > 86400000) return; // 24h过期
    if (state.selectedCol) selectedCol = state.selectedCol;
    if (state.selectedCells) state.selectedCells.forEach(k => selectedCells.add(k));
    if (state.selectedRows) state.selectedRows.forEach(k => selectedRows.add(k));
    if (state.searchTerm) { _searchTerm = state.searchTerm; document.getElementById("globalSearch").value = _searchTerm; }
    if (state.selectedCol || state.selectedCells.length || state.selectedRows.length) {
      document.getElementById("selText").innerHTML = state.selectedCol ? `列 <b>${escHtml(state.selectedCol)}</b>` : "已恢复选择";
    }
  } catch(e) {}
}

async function resetDefault() {
  await fetch("/reset", { method:"POST" });
  clearSelection(); loadTable();
  toast("已恢复默认数据", "ok");
}

// ===================== Toast =====================
function toast(msg, type) {
  const el = document.getElementById("toast");
  el.textContent = msg; el.className = "toast " + (type||"") + " show";
  clearTimeout(el._tid);
  el._tid = setTimeout(() => el.classList.remove("show"), 2200);
}

// ===================== 拖拽列排序 =====================
let _dragSrcCol = null;

function dragCol(e, col) {
  _dragSrcCol = col;
  e.dataTransfer.effectAllowed = "move";
  e.dataTransfer.setData("text/plain", col);
  e.target.style.opacity = "0.4";
}

function dragOver(e) {
  e.preventDefault();
  e.dataTransfer.dropEffect = "move";
}

function dragEnd(e) {
  e.target.style.opacity = "1";
}

async function dropCol(e, targetCol) {
  e.preventDefault();
  if (!_dragSrcCol || _dragSrcCol === targetCol) { _dragSrcCol = null; return; }
  const srcIdx = _cols.indexOf(_dragSrcCol);
  const tgtIdx = _cols.indexOf(targetCol);
  if (srcIdx < 0 || tgtIdx < 0) return;
  const newOrder = [..._cols];
  newOrder.splice(srcIdx, 1);
  newOrder.splice(tgtIdx, 0, _dragSrcCol);
  const res = await fetch("/process", {
    method:"POST",
    headers:{"Content-Type":"application/json"},
    body: JSON.stringify({ action:"reorder_cols", params:{order:newOrder} })
  });
  const j = await res.json();
  toast(j.message || "已重排", j.ok?"ok":"err");
  selectedCol = null;
  loadTable();
}

// ===================== 面板折叠 =====================
function togglePanel(h3) {
  h3.closest(".panel").classList.toggle("collapsed");
}

// ===================== 右键菜单 =====================
document.addEventListener("DOMContentLoaded", () => {
  loadTable();
  loadAutoSave();
  document.addEventListener("contextmenu", (e) => {
    const td = e.target.closest("td");
    if (!td || td.cellIndex <= 1) return;
    e.preventDefault();
    const menu = document.getElementById("ctxMenu");
    const row = td.parentElement.rowIndex;
    const colIdx = td.cellIndex - 2;
    const colName = _cols[colIdx] || "";
    const val = _rows[row]?.[colName] ?? "";
    menu.innerHTML = `
      <div class="item" onclick="selectCell(event,${row},'${escAttr(colName)}');hideCtx();loadTable();">📌 选中此单元格</div>
      <div class="item" onclick="selectCol('${escAttr(colName)}');hideCtx();loadTable();">📊 选中整列「${escHtml(colName)}」</div>
      <div class="item" onclick="startEdit(${row},'${escAttr(colName)}');hideCtx();">✏ 编辑此单元格</div>
      <div class="item" onclick="copyCell(${row},'${escAttr(colName)}');hideCtx();">📋 复制值 «${escHtml(String(val).substring(0,20))}»</div>
      <div class="item" onclick="openStatsForCol('${escAttr(colName)}');hideCtx();">📊 列统计「${escHtml(colName)}」</div>
    `;
    menu.style.display = "block";
    menu.style.left = (e.pageX + 2) + "px";
    menu.style.top = (e.pageY + 2) + "px";
  });
  document.addEventListener("click", (e) => {
    if (!e.target.closest(".ctx-menu")) document.getElementById("ctxMenu").style.display = "none";
  });

  // 拖拽结束恢复样式
  document.addEventListener("dragend", (e) => {
    if (e.target.tagName === "TH") e.target.style.opacity = "1";
  });

  // 键盘快捷键
  document.addEventListener("keydown", (e) => {
    if (e.target.tagName === "INPUT" || e.target.tagName === "TEXTAREA" || e.target.tagName === "SELECT") {
      // 在输入框内时，仅 Esc 退出编辑模式
      if (e.key === "Escape" && _editingCell) {
        e.preventDefault();
        _editingCell = null;
        loadTable();
      }
      return;
    }
    if ((e.ctrlKey || e.metaKey) && e.key === "z") {
      e.preventDefault(); undo();
    } else if ((e.ctrlKey || e.metaKey) && e.key === "y") {
      e.preventDefault(); redo();
    } else if ((e.ctrlKey || e.metaKey) && e.key === "s") {
      e.preventDefault(); saveCSV();
    } else if ((e.ctrlKey || e.metaKey) && e.key === "f") {
      e.preventDefault(); document.getElementById("globalSearch").focus();
    }
  });
});

function hideCtx() { document.getElementById("ctxMenu").style.display = "none"; }

async function copyCell(row, col) {
  const val = _rows[row]?.[col] ?? "";
  try { await navigator.clipboard.writeText(String(val)); toast("已复制", "ok"); }
  catch { toast("复制失败", "err"); }
}
</script>
</body>
</html>"""


# ---------------------------------------------------------------------------
# 路由
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    return render_template_string(INDEX_HTML)


@app.route("/api/data")
def api_data():
    with DATA_LOCK:
        df = DF.copy()

    search = request.args.get("search", "").strip()
    if search:
        try:
            re.compile(search)
            mask = pd.Series([False] * len(df))
            for c in df.columns:
                ser = df[c].astype(str)
                mask |= ser.str.contains(search, na=False, regex=True)
            df = df[mask].reset_index(drop=True)
        except re.error:
            mask = pd.Series([False] * len(df))
            for c in df.columns:
                ser = df[c].astype(str)
                mask |= ser.str.contains(search, na=False, regex=False)
            df = df[mask].reset_index(drop=True)

    cols = list(df.columns)
    rows = df.where(pd.notna(df), None).to_dict(orient="records")
    types = {}
    for c in cols:
        types[c] = _detect_type(df[c])

    # 条件格式: 数值列按分位数分 5 档
    cond_format = {}
    for c in cols:
        try:
            nums = pd.to_numeric(df[c], errors="coerce")
            if nums.notna().sum() >= 2:
                # 用分位数分档
                bins = [nums.min() - 1, nums.quantile(0.2), nums.quantile(0.4),
                        nums.quantile(0.6), nums.quantile(0.8), nums.max() + 1]
                labels = range(5)
                binned = pd.cut(nums, bins=bins, labels=labels, duplicates="drop")
                cond_format[c] = {i: int(v) for i, v in enumerate(binned) if pd.notna(v)}
        except Exception:
            pass

    can_undo = len(UNDO_STACK) > 0
    can_redo = len(REDO_STACK) > 0

    return jsonify({
        "columns": cols, "rows": rows, "types": types,
        "rowCount": len(df), "can_undo": can_undo, "can_redo": can_redo,
        "cond_format": cond_format
    })


@app.route("/upload", methods=["POST"])
def upload():
    global DF
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"ok": False, "message": "未选择文件"})
    try:
        fn = f.filename.lower()
        if fn.endswith(".csv"):
            raw = f.read().decode("utf-8-sig")
            new_df = pd.read_csv(io.StringIO(raw), dtype=str).fillna("")
        elif fn.endswith(".xlsx"):
            new_df = pd.read_excel(f, dtype=str).fillna("")
        else:
            return jsonify({"ok": False, "message": "仅支持 .csv 和 .xlsx 文件"})
        with DATA_LOCK:
            _push_history()
            _clear_redo()
            DF = new_df
        return jsonify({"ok": True, "message": f"已加载: {f.filename} ({len(DF)}行 x {len(DF.columns)}列)"})
    except Exception as e:
        return jsonify({"ok": False, "message": f"解析失败: {str(e)}"})


@app.route("/reset", methods=["POST"])
def reset():
    with DATA_LOCK:
        _push_history()
        _clear_redo()
        _reset_default()
    return jsonify({"ok": True, "message": "已恢复默认数据"})


@app.route("/save")
def save():
    with DATA_LOCK:
        buf = df_to_csv_bytes(DF)
    return send_file(buf, mimetype="text/csv; charset=utf-8",
                     as_attachment=True, download_name="acm_team_data.csv")


@app.route("/save_excel")
def save_excel():
    """导出为带样式的 Excel 文件"""
    with DATA_LOCK:
        df = DF.copy()

    wb = Workbook()
    ws = wb.active
    ws.title = "ACM 队伍数据"

    # 样式定义
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="00E5FF")
    header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    header_border = Border(bottom=Side(style="medium", color="00E5FF"))
    cell_font = Font(name="JetBrains Mono", size=10, color="E2E8F0")
    cell_fill = PatternFill(start_color="0A0E17", end_color="0A0E17", fill_type="solid")
    cell_border = Border(bottom=Side(style="thin", color="1E2D45"))
    center_align = Alignment(horizontal="center", vertical="center")

    # 写表头
    for ci, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=str(col_name))
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = center_align

    # 写数据
    for ri in range(len(df)):
        for ci, col_name in enumerate(df.columns, 1):
            val = df.iloc[ri, ci - 1]
            cell = ws.cell(row=ri + 2, column=ci, value=str(val) if pd.notna(val) else "")
            cell.font = cell_font
            cell.fill = cell_fill
            cell.border = cell_border
            cell.alignment = center_align

    # 列宽自适应
    for ci, col_name in enumerate(df.columns, 1):
        max_len = len(str(col_name))
        for ri in range(min(len(df), 100)):
            val = str(df.iloc[ri, ci - 1]) if pd.notna(df.iloc[ri, ci - 1]) else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="acm_team_data.xlsx")


# ---------------------------------------------------------------------------
# HTML 报告导出
# ---------------------------------------------------------------------------
@app.route("/save_html_report")
def save_html_report():
    """导出完整 HTML 数据分析报告"""
    with DATA_LOCK:
        df = DF.copy()

    cols = list(df.columns)
    rows = df.where(pd.notna(df), None).to_dict(orient="records")

    # 计算各列统计
    col_stats = {}
    for c in cols:
        ser = df[c].astype(str)
        nums = pd.to_numeric(ser, errors="coerce")
        valid_nums = nums.dropna()
        st = {
            "总数": int(len(ser)),
            "唯一值": int(ser.nunique()),
            "空值": int((ser == "").sum() + ser.isna().sum()),
        }
        if len(valid_nums) > 0:
            st["最小值"] = f"{valid_nums.min():.2f}"
            st["最大值"] = f"{valid_nums.max():.2f}"
            st["平均值"] = f"{valid_nums.mean():.2f}"
            st["中位数"] = f"{valid_nums.median():.2f}"
            st["标准差"] = f"{valid_nums.std():.2f}"
        col_stats[c] = st

    # 数据质量评分
    quality_score = 100
    for c in cols:
        ser = df[c].astype(str)
        empty_rate = ((ser == "").sum() + ser.isna().sum()) / len(ser)
        quality_score -= empty_rate * 20
    quality_score = max(0, min(100, int(quality_score)))

    # 构建列统计 HTML
    col_stats_html_parts = []
    for c in cols:
        stats_items = "".join(
            '<div class="stat-item"><div class="val">' + str(v) + '</div><div class="lbl">' + str(k) + '</div></div>'
            for k, v in col_stats[c].items()
        )
        col_stats_html_parts.append(
            '<details style="margin-bottom:6px;"><summary style="color:#00e5ff;cursor:pointer;font-weight:600;">'
            + c + ' (' + str(len(col_stats[c])) + '项指标)</summary>'
            + '<div class="stats-grid" style="margin-top:8px;">' + stats_items + '</div></details>'
        )

    # 数据预览表
    preview_rows_html = ""
    for i in range(min(50, len(rows))):
        row_html = "<tr>"
        for c in cols:
            row_html += "<td>" + (str(rows[i][c]) if i < len(rows) else "") + "</td>"
        row_html += "</tr>"
        preview_rows_html += row_html

    now_str = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')
    score_class = "score-high" if quality_score >= 80 else "score-mid"

    html = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8"><title>ACM 数据分析报告</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.7/dist/chart.umd.min.js"></script>
<style>
  * { margin:0; padding:0; box-sizing:border-box; }
  body { font-family:"Microsoft YaHei",sans-serif; background:#0a0e17; color:#e2e8f0; padding:20px; }
  h1 { color:#00e5ff; margin-bottom:10px; font-size:28px; }
  .meta { color:#94a3b8; margin-bottom:20px; font-size:13px; }
  .score { display:inline-block; padding:6px 18px; border-radius:20px; font-weight:700; font-size:18px; }
  .score-high { background:rgba(0,230,118,0.2); color:#00e676; }
  .score-mid { background:rgba(255,171,0,0.2); color:#ffab00; }
  .section { background:#111827; border:1px solid #1e2d45; border-radius:12px; padding:18px; margin-bottom:16px; }
  .section h2 { color:#00e5ff; font-size:18px; margin-bottom:12px; }
  table { width:100%; border-collapse:collapse; font-size:12px; margin-bottom:10px; }
  th { background:#1a2740; color:#00e5ff; padding:8px 12px; text-align:left; border-bottom:2px solid #00e5ff; }
  td { padding:7px 12px; border-bottom:1px solid rgba(30,45,69,0.6); }
  .stats-grid { display:grid; grid-template-columns:repeat(auto-fill, minmax(160px,1fr)); gap:8px; }
  .stat-item { background:#161e2c; padding:10px; border-radius:8px; text-align:center; border:1px solid #1e2d45; }
  .stat-item .val { color:#00e5ff; font-size:18px; font-weight:700; }
  .stat-item .lbl { color:#94a3b8; font-size:10px; }
  .chart-box { margin-top:16px; }
  canvas { max-height:300px; }
  @media print { body { background:#fff; color:#000; } .section { break-inside:avoid; border-color:#ccc; } }
</style></head>
<body>
<h1>ACM 队伍数据分析报告</h1>
<div class="meta">""" + f"生成于 {now_str} | {len(rows)} 行 × {len(cols)} 列 | 数据质量评分: <span class=\"score {score_class}\">{quality_score} 分</span>" + """</div>

<div class="section"><h2>列统计概览</h2>
""" + "".join(col_stats_html_parts) + """
</div>

<div class="section"><h2>数据预览 (前50行)</h2>
<table><thead><tr>""" + "".join("<th>" + c + "</th>" for c in cols) + """</tr></thead>
<tbody>""" + preview_rows_html + """</tbody></table></div>

<div class="section"><h2>可视化</h2>
<div class="chart-box"><canvas id="reportChart"></canvas></div></div>
<script>
(() => {
  const cols = """ + json.dumps(cols) + """;
  const rows = """ + json.dumps(rows[:50]) + """;
  var numCol = null;
  for (var ci = 0; ci < cols.length; ci++) {
    var v = rows[0] && rows[0][cols[ci]];
    if (v && !isNaN(parseFloat(v))) { numCol = cols[ci]; break; }
  }
  if (numCol) {
    new Chart(document.getElementById('reportChart'), {
      type: 'bar',
      data: {
        labels: rows.map(function(r) { return r[cols[1]] || r[cols[0]] || ''; }),
        datasets: [{
          label: numCol, data: rows.map(function(r) { return parseFloat(r[numCol]) || 0; }),
          backgroundColor: 'rgba(0,229,255,0.3)', borderColor:'#00e5ff', borderWidth:1
        }]
      },
      options: { responsive:true, plugins:{legend:{labels:{color:'#94a3b8'}}}, scales:{ y:{beginAtZero:true,grid:{color:'rgba(30,45,69,0.5)'},ticks:{color:'#94a3b8'}}, x:{ticks:{color:'#94a3b8'}} } }
    });
  }
})();
</script>
</body></html>"""
    buf = io.BytesIO()
    buf.write(html.encode("utf-8"))
    buf.seek(0)
    return send_file(buf, mimetype="text/html; charset=utf-8",
                     as_attachment=True, download_name="acm_report.html")


# ---------------------------------------------------------------------------
# AI 数据洞察
# ---------------------------------------------------------------------------
@app.route("/api/insights")
def api_insights():
    """自动生成数据洞察"""
    with DATA_LOCK:
        df = DF.copy()

    insights = []
    cols = list(df.columns)
    total_rows = len(df)

    # 整体概览
    insights.append({"type": "overview", "title": "数据概览",
                     "text": f"数据集包含 {total_rows} 条记录，{len(cols)} 个字段，数据完整度良好。"})

    # 列类型分布
    num_cols = []
    text_cols = []
    for c in cols:
        try:
            pd.to_numeric(df[c])
            num_cols.append(c)
        except (ValueError, TypeError):
            text_cols.append(c)

    if num_cols:
        insights.append({"type": "columns", "title": "数值列分析",
                         "text": f"发现 {len(num_cols)} 个数值列: {', '.join(num_cols)}，可进行统计分析和可视化。"})

    # 每列洞察
    for c in cols:
        ser = df[c].astype(str)
        unique_count = ser.nunique()
        empty_count = int((ser == "").sum() + ser.isna().sum())

        # 高唯一值率
        if unique_count == total_rows:
            insights.append({"type": "info", "title": f"「{c}」全部唯一",
                             "text": f"列「{c}」的 {total_rows} 个值全部不同，可能是主键或标识列。"})

        # 高空值率警告
        if empty_count > total_rows * 0.3:
            insights.append({"type": "warning", "title": f"「{c}」空值较多",
                             "text": f"列「{c}」有 {empty_count} 个空值 ({empty_count / total_rows * 100:.1f}%)，建议清理或填充。"})

        # 低唯一值率 (可能是分类列)
        if unique_count <= 5 and unique_count < total_rows:
            top_vals = ser.value_counts().head(3)
            insights.append({"type": "category", "title": f"「{c}」分类列",
                             "text": f"列「{c}」仅有 {unique_count} 个不同值，分布: {', '.join(f'{k}:{v}次' for k,v in top_vals.items())}"})

        # 数值列统计
        try:
            nums = pd.to_numeric(ser, errors="coerce")
            valid = nums.dropna()
            if len(valid) >= 3:
                mean_val = valid.mean()
                std_val = valid.std()
                cv = std_val / mean_val * 100 if mean_val != 0 else 0

                # 异常值检测 (IQR)
                q1, q3 = valid.quantile(0.25), valid.quantile(0.75)
                iqr = q3 - q1
                outliers = valid[(valid < q1 - 1.5 * iqr) | (valid > q3 + 1.5 * iqr)]
                if len(outliers) > 0:
                    insights.append({"type": "warning", "title": f"「{c}」异常值检测",
                                     "text": f"发现 {len(outliers)} 个可能的异常值 (IQR方法): {', '.join(f'{x:.1f}' for x in outliers.head(5).tolist())}{'...' if len(outliers) > 5 else ''}"})

                # 变异系数大 - 数据分散
                if cv > 50:
                    insights.append({"type": "info", "title": f"「{c}」数据分散",
                                     "text": f"变异系数 {cv:.1f}%，数据分布较为分散 (均值 {mean_val:.1f} ± {std_val:.1f})。"})

                # 偏态检测
                skew = valid.skew()
                if abs(skew) > 1:
                    direction = "右偏" if skew > 0 else "左偏"
                    insights.append({"type": "info", "title": f"「{c}」{direction}分布",
                                     "text": f"偏度 {skew:.2f}，数据呈{direction}分布，中位数 ({valid.median():.1f}) 与均值 ({mean_val:.1f}) 存在差异。"})
        except Exception:
            pass

    # 列间相关性 (Pearson)
    if len(num_cols) >= 2:
        try:
            num_df = df[num_cols].apply(pd.to_numeric, errors="coerce")
            corr = num_df.corr()
            for i in range(len(num_cols)):
                for j in range(i + 1, len(num_cols)):
                    r = corr.iloc[i, j]
                    if pd.notna(r) and abs(r) > 0.5:
                        direction = "正相关" if r > 0 else "负相关"
                        insights.append({"type": "correlation", "title": f"「{num_cols[i]}」与「{num_cols[j]}」{direction}",
                                         "text": f"相关系数 r = {r:.3f}，两者存在显著{direction}关系。"})
        except Exception:
            pass

    # 数据质量评分
    quality_score = 100
    for c in cols:
        ser = df[c].astype(str)
        empty_rate = ((ser == "").sum() + ser.isna().sum()) / max(len(ser), 1)
        quality_score -= empty_rate * 20
    quality_score = max(0, min(100, int(quality_score)))

    # 重复行
    dup_count = int(df.duplicated().sum())
    if dup_count > 0:
        insights.append({"type": "warning", "title": "重复数据",
                         "text": f"发现 {dup_count} 行重复数据，建议去重处理。"})

    return jsonify({
        "ok": True,
        "quality_score": quality_score,
        "quality_grade": "优秀" if quality_score >= 90 else "良好" if quality_score >= 75 else "一般" if quality_score >= 60 else "较差",
        "total_rows": total_rows,
        "total_cols": len(cols),
        "num_cols": len(num_cols),
        "text_cols_count": len(text_cols),
        "insights": insights
    })


# ---------------------------------------------------------------------------
# 合并数据集
# ---------------------------------------------------------------------------
@app.route("/api/merge", methods=["POST"])
def api_merge():
    """合并上传的 CSV/Excel 与当前数据"""
    global DF
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"ok": False, "message": "未选择文件"})

    try:
        fn = f.filename.lower()
        if fn.endswith(".csv"):
            raw = f.read().decode("utf-8-sig")
            new_df = pd.read_csv(io.StringIO(raw), dtype=str).fillna("")
        elif fn.endswith(".xlsx"):
            new_df = pd.read_excel(f, dtype=str).fillna("")
        else:
            return jsonify({"ok": False, "message": "仅支持 .csv 和 .xlsx 文件"})

        mode = request.form.get("mode", "concat")
        with DATA_LOCK:
            _push_history()
            _clear_redo()
            if mode == "concat":
                DF = pd.concat([DF, new_df], ignore_index=True)
                msg = f"纵向合并完成 — {len(DF)} 行 × {len(DF.columns)} 列"
            elif mode == "join":
                join_col = request.form.get("join_col", DF.columns[0])
                if join_col in DF.columns and join_col in new_df.columns:
                    DF = pd.merge(DF, new_df, on=join_col, how="outer")
                    msg = f"按「{join_col}」横向合并完成 — {len(DF)} 行 × {len(DF.columns)} 列"
                else:
                    return jsonify({"ok": False, "message": f"列「{join_col}」在两个数据集中必须都存在"})
            else:
                return jsonify({"ok": False, "message": f"未知合并模式: {mode}"})
        return jsonify({"ok": True, "message": msg})
    except Exception as e:
        return jsonify({"ok": False, "message": f"合并失败: {str(e)}"})


# ---------------------------------------------------------------------------
# 内联编辑接口
# ---------------------------------------------------------------------------
@app.route("/edit_cell", methods=["POST"])
def edit_cell():
    data = request.get_json(force=True)
    row = data.get("row")
    col = data.get("col")
    value = data.get("value", "")
    with DATA_LOCK:
        global DF
        if row is None or col is None or col not in DF.columns:
            return jsonify({"ok": False, "message": "无效的单元格"})
        if row < 0 or row >= len(DF):
            return jsonify({"ok": False, "message": "行索引越界"})
        try:
            _push_history()
            _clear_redo()
            DF.at[row, col] = value
            return jsonify({"ok": True, "message": "已保存"})
        except Exception as e:
            return jsonify({"ok": False, "message": str(e)})


# ---------------------------------------------------------------------------
# 撤销 / 重做
# ---------------------------------------------------------------------------
@app.route("/api/undo", methods=["POST"])
def api_undo():
    global DF
    with DATA_LOCK:
        if not UNDO_STACK:
            return jsonify({"ok": False, "message": "没有可撤销的操作"})
        REDO_STACK.append(DF.copy())
        DF = UNDO_STACK.pop()
        return jsonify({"ok": True, "message": f"已撤销 (剩余 {len(UNDO_STACK)} 步)"})


@app.route("/api/redo", methods=["POST"])
def api_redo():
    global DF
    with DATA_LOCK:
        if not REDO_STACK:
            return jsonify({"ok": False, "message": "没有可重做的操作"})
        _push_history()  # 当前状态进 undo 栈
        DF = REDO_STACK.pop()
        return jsonify({"ok": True, "message": f"已重做 (剩余 {len(REDO_STACK)} 步)"})


# ---------------------------------------------------------------------------
# 行操作
# ---------------------------------------------------------------------------
@app.route("/api/row/add", methods=["POST"])
def api_add_row():
    global DF
    with DATA_LOCK:
        _push_history()
        _clear_redo()
        empty_row = {c: "" for c in DF.columns}
        DF = pd.concat([DF, pd.DataFrame([empty_row])], ignore_index=True)
        return jsonify({"ok": True, "message": f"已添加行，共 {len(DF)} 行"})


@app.route("/api/row/delete", methods=["POST"])
def api_delete_rows():
    global DF
    data = request.get_json(force=True)
    rows = sorted(set(data.get("rows", [])), reverse=True)
    if not rows:
        return jsonify({"ok": False, "message": "未指定行"})
    with DATA_LOCK:
        _push_history()
        _clear_redo()
        DF = DF.drop(rows).reset_index(drop=True)
        return jsonify({"ok": True, "message": f"已删除 {len(rows)} 行，剩余 {len(DF)} 行"})


@app.route("/api/row/duplicate", methods=["POST"])
def api_duplicate_rows():
    global DF
    data = request.get_json(force=True)
    rows = sorted(data.get("rows", []))
    if not rows:
        return jsonify({"ok": False, "message": "未指定行"})
    with DATA_LOCK:
        _push_history()
        _clear_redo()
        copies = DF.iloc[rows].copy()
        DF = pd.concat([DF, copies], ignore_index=True)
        return jsonify({"ok": True, "message": f"已复制 {len(rows)} 行，共 {len(DF)} 行"})


# ---------------------------------------------------------------------------
# 列统计
# ---------------------------------------------------------------------------
@app.route("/api/stats/<col>")
def api_col_stats(col):
    with DATA_LOCK:
        df = DF.copy()

    if col not in df.columns:
        return jsonify({"ok": False, "message": f"列「{col}」不存在"})

    series = df[col].astype(str)
    total = int(len(series))
    unique = int(series.nunique())
    empty = int(series.isna().sum() + (series == "").sum())
    stats = {
        "总数": total,
        "唯一值": unique,
        "空值数": empty,
        "非空率": f"{(total - empty) / total * 100:.1f}%" if total > 0 else "0%",
    }

    # 数值型统计
    try:
        nums = pd.to_numeric(series, errors="coerce")
        valid = nums.dropna()
        if len(valid) > 0:
            stats["最小值"] = f"{valid.min():.2f}"
            stats["最大值"] = f"{valid.max():.2f}"
            stats["平均值"] = f"{valid.mean():.2f}"
            stats["中位数"] = f"{valid.median():.2f}"
            stats["标准差"] = f"{valid.std():.2f}"
            stats["总和"] = f"{valid.sum():.2f}"
    except Exception:
        pass

    # 文本型统计
    stats["最长值"] = str(series.str.len().max())
    stats["最短值"] = str(series.str.len().min())
    stats["平均长度"] = f"{series.str.len().mean():.1f}"

    # 直方图数据 (数值列取分段，文本列取值分布)
    histogram = {"labels": [], "values": []}
    try:
        nums = pd.to_numeric(series, errors="coerce")
        valid = nums.dropna()
        if len(valid) >= 2:
            bins = 8
            counts, bin_edges = np.histogram(valid, bins=bins)
            histogram["labels"] = [f"{bin_edges[i]:.1f}-{bin_edges[i+1]:.1f}" for i in range(len(counts))]
            histogram["values"] = counts.tolist()
        else:
            # 文本分布: top 10 值
            top = series.value_counts().head(10)
            histogram["labels"] = top.index.tolist()
            histogram["values"] = top.values.tolist()
    except Exception:
        top = series.value_counts().head(10)
        histogram["labels"] = top.index.tolist()
        histogram["values"] = top.values.tolist()

    return jsonify({"ok": True, "stats": stats, "histogram": histogram})


# ---------------------------------------------------------------------------
# 批量处理
# ---------------------------------------------------------------------------
def _apply_cell_op(action: str, params: dict, val: str) -> str:
    if action == "delete_chars":
        for ch in params.get("chars", ""):
            val = val.replace(ch, "")
    elif action == "replace":
        val = val.replace(params.get("old", ""), params.get("new", ""))
    elif action == "case":
        mode = params.get("mode", "upper")
        val = {"upper": val.upper(), "lower": val.lower(),
               "title": val.title(), "capitalize": val.capitalize()}.get(mode, val)
    elif action == "reverse":
        val = val[::-1]
    elif action == "strip":
        mode = params.get("mode", "both")
        val = {"both": val.strip(), "left": val.lstrip(),
               "right": val.rstrip(), "all": val.replace(" ", "")}.get(mode, val)
    elif action == "prefix":
        val = params.get("text", "") + val
    elif action == "suffix":
        val = val + params.get("text", "")
    elif action == "slice":
        s, e = params.get("start", ""), params.get("end", "")
        si = int(s) if (s.lstrip("-").isdigit() and s.strip()) else None
        ei = int(e) if (e.lstrip("-").isdigit() and e.strip()) else None
        val = val[si:ei]
    elif action == "regex":
        pat, repl = params.get("pattern", ""), params.get("repl", "")
        if pat: val = re.sub(pat, repl, val)
    elif action == "split_take":
        sep = params.get("sep", "")
        if sep:
            parts = val.split(sep)
            try:
                idx = int(params.get("idx", "0"))
                val = parts[idx] if 0 <= idx < len(parts) else val
            except (ValueError, IndexError):
                pass
    elif action == "extract_digits":
        val = "".join(ch for ch in val if ch.isdigit())
    elif action == "extract_letters":
        val = "".join(ch for ch in val if ch.isalpha())
    elif action == "extract_chinese":
        val = "".join(ch for ch in val if _is_chinese(ch))
    elif action == "remove_digits":
        val = "".join(ch for ch in val if not ch.isdigit())
    elif action == "remove_letters":
        val = "".join(ch for ch in val if not ch.isalpha())
    elif action == "remove_chinese":
        val = "".join(ch for ch in val if not _is_chinese(ch))
    elif action == "remove_punct":
        val = re.sub(r"[^\w\s一-鿿㐀-䶿豈-﫿]", "", val)
    elif action == "pad_left":
        w, ch = int(params.get("width", 0) or 0), params.get("char", " ") or " "
        val = val.rjust(max(w, len(val)), ch)
    elif action == "pad_right":
        w, ch = int(params.get("width", 0) or 0), params.get("char", " ") or " "
        val = val.ljust(max(w, len(val)), ch)
    elif action == "to_number":
        try:
            v = pd.to_numeric(val)
            val = str(v)
        except Exception:
            pass
    elif action == "to_text":
        val = str(val)
    elif action == "clear":
        val = ""
    return val


@app.route("/process", methods=["POST"])
def process():
    data = request.get_json(force=True)
    action = data.get("action", "")
    col = data.get("col", None)
    cells = data.get("cells", None)
    params = data.get("params", {})

    with DATA_LOCK:
        global DF
        df = DF
        msg, ok = "操作完成", True

        try:
            # 列操作
            if action == "delete":
                if col and col in df.columns:
                    _push_history(); _clear_redo()
                    DF = df.drop(columns=[col]); msg = f"已删除列「{col}」"
            elif action == "sort_asc":
                if col and col in df.columns:
                    _push_history(); _clear_redo()
                    DF = _smart_sort(df, col, ascending=True)
                    msg = f"已按「{col}」升序排列"
            elif action == "sort_desc":
                if col and col in df.columns:
                    _push_history(); _clear_redo()
                    DF = _smart_sort(df, col, ascending=False)
                    msg = f"已按「{col}」降序排列"
            elif action == "dedup":
                if col and col in df.columns:
                    _push_history(); _clear_redo()
                    DF = df.drop_duplicates(subset=[col]).reset_index(drop=True)
                    msg = f"已按「{col}」去重"
            elif action == "count_values":
                if col and col in df.columns:
                    _push_history(); _clear_redo()
                    cnt = df[col].value_counts().reset_index()
                    cnt.columns = [col, "次数"]
                    DF = cnt; msg = f"「{col}」值计数完成"
            elif action == "fill_empty":
                if col and col in df.columns:
                    _push_history(); _clear_redo()
                    DF[col] = df[col].fillna("").replace("", params.get("value", ""))
                    msg = f"列「{col}」空值已填充"
            elif action == "clone":
                if col and col in df.columns:
                    _push_history(); _clear_redo()
                    col_str, new_name = str(col), str(col) + "_副本"
                    idx = list(df.columns).index(col_str) + 1
                    DF = pd.concat([df.iloc[:, :idx], df[[col_str]].rename(columns={col_str: new_name}), df.iloc[:, idx:]], axis=1)
                    msg = f"已克隆列「{col}」→「{new_name}」"
            elif action == "rename":
                new_name = str(params.get("new_name", ""))
                if col and col in df.columns and new_name and new_name != col:
                    if new_name in df.columns:
                        ok, msg = False, f"列名「{new_name}」已存在"
                    else:
                        _push_history(); _clear_redo()
                        DF = df.rename(columns={str(col): new_name}); msg = f"已重命名「{col}」→「{new_name}」"
            # 筛选
            elif action == "filter":
                if col and col in df.columns:
                    _push_history(); _clear_redo()
                    op, val = params.get("op", "contains"), params.get("val", "")
                    ser = df[col].astype(str)
                    masks = {
                        "contains": ser.str.contains(val, na=False, regex=False),
                        "equals": ser == val,
                        "startswith": ser.str.startswith(val, na=False),
                        "endswith": ser.str.endswith(val, na=False),
                        "regex": ser.str.contains(val, na=False, regex=True),
                        "gt": pd.to_numeric(ser, errors="coerce") > float(val),
                        "lt": pd.to_numeric(ser, errors="coerce") < float(val),
                        "not_empty": ser.notna() & (ser != ""),
                        "is_empty": ser.isna() | (ser == ""),
                    }
                    mask = masks.get(op, pd.Series([True] * len(df)))
                    DF = df[mask].reset_index(drop=True)
                    msg = f"筛选后保留 {len(DF)} 行"
            elif action == "clear_filter":
                _push_history(); _clear_redo()
                _reset_default(); msg = "已恢复默认数据"
            # 计算列
            elif action == "calc_col":
                src_col, new_col = col, params.get("new_col", "计算结果")
                op = params.get("op", "len")
                if src_col and src_col in df.columns and new_col not in df.columns:
                    _push_history(); _clear_redo()
                    ser = df[src_col].astype(str)
                    if op == "len": DF[new_col] = ser.str.len()
                    elif op == "word_count": DF[new_col] = ser.str.split().str.len()
                    elif op == "is_palindrome": DF[new_col] = ser.apply(lambda x: "是" if x == x[::-1] else "否")
                    elif op == "digit_sum": DF[new_col] = ser.apply(lambda x: sum(int(ch) for ch in x if ch.isdigit()))
                    elif op == "upper": DF[new_col] = ser.str.upper()
                    elif op == "lower": DF[new_col] = ser.str.lower()
                    elif op == "reverse": DF[new_col] = ser.apply(lambda x: x[::-1])
                    elif op == "strip": DF[new_col] = ser.str.strip()
                    msg = f"已添加计算列「{new_col}」"
                elif new_col in df.columns:
                    ok, msg = False, f"列名「{new_col}」已存在"
            # 透视表
            elif action == "pivot":
                row_col = col
                val_col = params.get("val_col", "")
                agg = params.get("agg", "count")
                if row_col and val_col and row_col in df.columns and val_col in df.columns:
                    _push_history(); _clear_redo()
                    if agg == "count":
                        pivot = df.groupby(row_col)[val_col].count().reset_index()
                    elif agg == "sum":
                        pivot = df.groupby(row_col)[val_col].apply(
                            lambda x: pd.to_numeric(x, errors="coerce").sum()
                        ).reset_index()
                    elif agg == "mean":
                        pivot = df.groupby(row_col)[val_col].apply(
                            lambda x: pd.to_numeric(x, errors="coerce").mean()
                        ).reset_index()
                    elif agg == "max":
                        pivot = df.groupby(row_col)[val_col].apply(
                            lambda x: pd.to_numeric(x, errors="coerce").max()
                        ).reset_index()
                    elif agg == "min":
                        pivot = df.groupby(row_col)[val_col].apply(
                            lambda x: pd.to_numeric(x, errors="coerce").min()
                        ).reset_index()
                    pivot.columns = [row_col, f"{agg}({val_col})"]
                    DF = pivot
                    msg = f"透视表完成 — {len(DF)} 行"
                else:
                    ok, msg = False, "透视参数无效"
            # 自定义公式列
            elif action == "formula_col":
                src_col = col
                new_col = params.get("new_col", "公式结果")
                formula = params.get("formula", "")
                if src_col and new_col not in df.columns and formula:
                    _push_history(); _clear_redo()
                    # 支持 {x} 作为当前单元格值, {idx} 作为行号
                    def eval_formula(val, idx):
                        try:
                            x = pd.to_numeric(val) if val else 0
                            return eval(formula, {"__builtins__": {}}, {"x": x, "idx": idx, "abs": abs, "round": round, "len": len, "str": str, "int": int, "float": float, "max": max, "min": min, "sum": sum})
                        except Exception:
                            return val
                    ser = df[src_col]
                    DF[new_col] = [eval_formula(v, i) for i, v in enumerate(ser)]
                    msg = f"已添加公式列「{new_col}」(公式: {formula})"
                elif new_col in df.columns:
                    ok, msg = False, f"列名「{new_col}」已存在"
            # 列重排
            elif action == "reorder_cols":
                new_order = params.get("order", [])
                if new_order and set(new_order) == set(df.columns):
                    _push_history(); _clear_redo()
                    DF = df[new_order]
                    msg = "列顺序已更新"
                else:
                    ok, msg = False, "列顺序不完整"
            # 数据脱敏
            elif action == "mask_col":
                if col and col in df.columns:
                    _push_history(); _clear_redo()
                    mode = params.get("mode", "partial")
                    ser = df[col].astype(str)
                    if mode == "partial":
                        # 保留首尾，中间用*替代
                        DF[col] = ser.apply(lambda x: x[0] + "*" * max(0, len(x) - 2) + x[-1] if len(x) > 2 else "*" * len(x))
                    elif mode == "stars":
                        DF[col] = "***"
                    elif mode == "hash":
                        import hashlib
                        DF[col] = ser.apply(lambda x: hashlib.md5(x.encode()).hexdigest()[:8])
                    msg = f"列「{col}」已脱敏处理"
            # 单元格批量操作
            elif cells:
                _push_history(); _clear_redo()
                df = DF.copy()
                for r, c in cells:
                    if r < 0 or r >= len(df) or c not in df.columns:
                        continue
                    df.at[r, c] = _apply_cell_op(action, params, str(df.at[r, c]))
                DF = df
                msg = f"已处理 {len(cells)} 个单元格"
            else:
                ok, msg = False, f"未知操作: {action}"
        except Exception as e:
            ok, msg = False, f"处理出错: {str(e)}"

    return jsonify({"ok": ok, "message": msg})


# ---------------------------------------------------------------------------
# SQLite 持久化
# ---------------------------------------------------------------------------
import sqlite3 as _sqlite3

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "acm_data.db")


def save_to_sqlite():
    """将 DataFrame 保存到 SQLite"""
    with _sqlite3.connect(DB_PATH) as conn:
        DF.to_sql("acm_data", conn, if_exists="replace", index=False)


def load_from_sqlite():
    """从 SQLite 加载 DataFrame"""
    global DF
    if not os.path.exists(DB_PATH):
        return False
    try:
        with _sqlite3.connect(DB_PATH) as conn:
            DF = pd.read_sql("SELECT * FROM acm_data", conn, dtype=str).fillna("")
        return True
    except Exception:
        return False


@app.route("/api/db/save", methods=["POST"])
def api_db_save():
    with DATA_LOCK:
        save_to_sqlite()
    return jsonify({"ok": True, "message": "已保存到数据库"})


@app.route("/api/db/load", methods=["POST"])
def api_db_load():
    global DF
    with DATA_LOCK:
        _push_history()
        _clear_redo()
        if load_from_sqlite():
            return jsonify({"ok": True, "message": f"已从数据库加载 ({len(DF)}行 x {len(DF.columns)}列)"})
        return jsonify({"ok": False, "message": "数据库中没有数据"})


# ---------------------------------------------------------------------------
# 备份系统
# ---------------------------------------------------------------------------
BACKUP_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backups")
os.makedirs(BACKUP_DIR, exist_ok=True)


@app.route("/api/backup/list")
def api_backup_list():
    backups = []
    for fn in sorted(os.listdir(BACKUP_DIR)):
        if fn.endswith(".csv"):
            fpath = os.path.join(BACKUP_DIR, fn)
            name = fn[:-4]
            size = os.path.getsize(fpath)
            mtime = os.path.getmtime(fpath)
            backups.append({
                "name": name,
                "size": f"{size / 1024:.1f}KB",
                "modified": pd.Timestamp(mtime, unit="s").strftime("%Y-%m-%d %H:%M")
            })
    return jsonify({"ok": True, "backups": backups})


@app.route("/api/backup/create", methods=["POST"])
def api_backup_create():
    data = request.get_json(force=True) or {}
    name = data.get("name", "").strip()
    if not name:
        name = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
    # sanitize filename
    name = re.sub(r"[^\w一-鿿_\-]", "_", name)
    fpath = os.path.join(BACKUP_DIR, name + ".csv")
    with DATA_LOCK:
        DF.to_csv(fpath, index=False, encoding="utf-8-sig")
    return jsonify({"ok": True, "message": f"备份已创建: {name}", "name": name})


@app.route("/api/backup/restore", methods=["POST"])
def api_backup_restore():
    global DF
    data = request.get_json(force=True)
    name = data.get("name", "").strip()
    name = re.sub(r"[^\w一-鿿_\-]", "_", name)
    fpath = os.path.join(BACKUP_DIR, name + ".csv")
    if not os.path.exists(fpath):
        return jsonify({"ok": False, "message": f"备份「{name}」不存在"})
    with DATA_LOCK:
        _push_history()
        _clear_redo()
        DF = pd.read_csv(fpath, dtype=str).fillna("")
    return jsonify({"ok": True, "message": f"已恢复备份「{name}」({len(DF)}行 x {len(DF.columns)}列)"})


@app.route("/api/backup/delete", methods=["POST"])
def api_backup_delete():
    data = request.get_json(force=True)
    name = data.get("name", "").strip()
    name = re.sub(r"[^\w一-鿿_\-]", "_", name)
    fpath = os.path.join(BACKUP_DIR, name + ".csv")
    if not os.path.exists(fpath):
        return jsonify({"ok": False, "message": f"备份「{name}」不存在"})
    os.remove(fpath)
    return jsonify({"ok": True, "message": f"已删除备份「{name}」"})


# ---------------------------------------------------------------------------
# 数据对比 (Diff)
# ---------------------------------------------------------------------------
@app.route("/api/diff", methods=["POST"])
def api_diff():
    """对比当前数据与上传文件或指定备份"""
    with DATA_LOCK:
        current = DF.copy()

    source = None
    source_name = ""

    # 从上传文件
    f = request.files.get("file")
    if f and f.filename:
        try:
            fn = f.filename.lower()
            if fn.endswith(".csv"):
                raw = f.read().decode("utf-8-sig")
                source = pd.read_csv(io.StringIO(raw), dtype=str).fillna("")
            elif fn.endswith(".xlsx"):
                source = pd.read_excel(f, dtype=str).fillna("")
            source_name = f.filename
        except Exception:
            pass

    # 从备份 (支持 query 参数或 form 参数)
    if source is None:
        backup_name = request.args.get("backup", "") or (request.form.get("backup", "") if request.form else "")
        if backup_name:
            backup_name = re.sub(r"[^\w一-鿿_\-]", "_", backup_name)
            fpath = os.path.join(BACKUP_DIR, backup_name + ".csv")
            if os.path.exists(fpath):
                source = pd.read_csv(fpath, dtype=str).fillna("")
                source_name = backup_name

    if source is None:
        return jsonify({"ok": False, "message": "请上传对比文件或指定备份名"})

    # 计算差异
    common_cols = [c for c in current.columns if c in source.columns]
    if not common_cols:
        return jsonify({"ok": False, "message": "两个数据集没有共同列"})

    # 按共同列合并找差异
    merged = pd.merge(current[common_cols], source[common_cols],
                      on=common_cols[0] if common_cols else None,
                      how="outer", indicator=True)
    only_current = int((merged["_merge"] == "left_only").sum())
    only_source = int((merged["_merge"] == "right_only").sum())
    in_both = int((merged["_merge"] == "both").sum())

    # 修改检测 (在共同行中)
    modified = 0
    if in_both > 0 and len(common_cols) > 1:
        both_mask = merged["_merge"] == "both"
        current_subset = current[common_cols]
        source_subset = source[common_cols]
        # 采样对比
        for col in common_cols[1:]:
            if col in current.columns and col in source.columns:
                cur_vals = set(current[col].dropna().astype(str).values)
                src_vals = set(source[col].dropna().astype(str).values)
                diff_vals = cur_vals.symmetric_difference(src_vals)
                if diff_vals:
                    modified += len(diff_vals)

    result = {
        "ok": True,
        "source": source_name,
        "current_rows": len(current),
        "source_rows": len(source),
        "common_cols": len(common_cols),
        "only_in_current": only_current,
        "only_in_source": only_source,
        "in_both": in_both,
        "modified_values": min(modified, 999),
        "summary": f"对比「{source_name}」: 当前 {len(current)}行 vs 来源 {len(source)}行 | 仅当前: {only_current} | 仅来源: {only_source} | 共同: {in_both}"
    }
    return jsonify(result)


# ---------------------------------------------------------------------------
# 自然语言查询
# ---------------------------------------------------------------------------
@app.route("/api/nlquery")
def api_nlquery():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify({"ok": False, "message": "请输入查询语句"})

    with DATA_LOCK:
        df = DF.copy()

    # 简单 NL 解析
    q_lower = q.lower()
    result_df = df
    explanation = []

    # 模式1: "列名 大于/小于/等于/包含 值"
    for col in df.columns:
        col_lower = col.lower()
        if col_lower in q_lower or col in q:
            # 数值比较
            m = re.search(rf'{re.escape(col)}[\s]*([大于小于等于>=<]+)[\s]*(\d+\.?\d*)', q)
            if m:
                op_str, val_str = m.group(1), m.group(2)
                val = float(val_str)
                nums = pd.to_numeric(df[col], errors="coerce")
                if "大于" in op_str or op_str == ">":
                    result_df = result_df[nums > val]
                    explanation.append(f"{col} > {val}")
                elif "小于" in op_str or op_str == "<":
                    result_df = result_df[nums < val]
                    explanation.append(f"{col} < {val}")
                elif "等于" in op_str or op_str == "=":
                    result_df = result_df[nums == val]
                    explanation.append(f"{col} = {val}")

            # 包含
            m2 = re.search(rf'{re.escape(col)}[\s]*(?:包含|含有|包括)[\s]*["\']?(.+?)["\']?(?:$|\s)', q)
            if not m2:
                m2 = re.search(rf'(?:包含|含有|包括)[\s]*["\']?(.+?)["\']?[\s]*(?:的|在){re.escape(col)}', q)
            if m2:
                keyword = m2.group(1).strip()
                result_df = result_df[result_df[col].astype(str).str.contains(keyword, na=False)]
                explanation.append(f"{col} 包含 '{keyword}'")

    # 模式2: "前N行" / "后N行" / "随机N行"
    m = re.search(r'[前后随机](\d+)行', q)
    if m:
        n = min(int(m.group(1)), len(result_df))
        if "前" in q:
            result_df = result_df.head(n)
            explanation.append(f"取前{n}行")
        elif "后" in q or "最后" in q:
            result_df = result_df.tail(n)
            explanation.append(f"取后{n}行")
        elif "随机" in q:
            result_df = result_df.sample(n=n)
            explanation.append(f"随机{n}行")

    # 模式3: "按X排序" / "按X降序"
    m = re.search(r'按(.+?)(降序|升序|排序)', q)
    if m:
        sort_col = m.group(1).strip()
        ascending = "降" not in m.group(2)
        if sort_col in df.columns:
            result_df = _smart_sort(result_df, sort_col, ascending=ascending)
            explanation.append(f"按{sort_col}{'升序' if ascending else '降序'}排列")

    # 模式4: "X最多的Y" -> groupby X, sum Y, top
    m = re.search(r'(.+?)最[多高](?:的)(.+)', q)
    if m:
        group_col = m.group(1).strip()
        val_col = m.group(2).strip()
        if group_col in df.columns and val_col in df.columns:
            try:
                agg = result_df.groupby(group_col)[val_col].apply(
                    lambda x: pd.to_numeric(x, errors="coerce").sum()
                ).sort_values(ascending=False).reset_index()
                agg.columns = [group_col, f"{val_col}_合计"]
                result_df = agg.head(5)
                explanation.append(f"{group_col} 按 {val_col} 合计排名")
            except Exception:
                pass

    result_df = result_df.reset_index(drop=True)

    # 限制返回行数
    if len(result_df) > 100:
        result_df = result_df.head(100)
        explanation.append("(仅显示前100行)")

    if not explanation:
        # 默认: 全文搜索
        mask = pd.Series([False] * len(result_df))
        for c in result_df.columns:
            ser = result_df[c].astype(str)
            mask |= ser.str.contains(q, na=False, regex=False)
        result_df = result_df[mask].reset_index(drop=True)
        explanation.append(f"全文字段匹配 '{q}'")

    cols = list(result_df.columns)
    rows = result_df.where(pd.notna(result_df), None).to_dict(orient="records")

    return jsonify({
        "ok": True,
        "query": q,
        "explanation": " → ".join(explanation),
        "result_count": len(result_df),
        "columns": cols,
        "rows": rows
    })


# ---------------------------------------------------------------------------
# 批量列操作
# ---------------------------------------------------------------------------
@app.route("/api/bulk_cols", methods=["POST"])
def api_bulk_cols():
    """对多个列同时执行操作"""
    data = request.get_json(force=True)
    cols = data.get("cols", [])
    action = data.get("action", "")
    params = data.get("params", {})

    if not cols:
        return jsonify({"ok": False, "message": "请选择要操作的列"})

    with DATA_LOCK:
        global DF
        _push_history()
        _clear_redo()
        msg = ""

        try:
            if action == "delete":
                DF = DF.drop(columns=[c for c in cols if c in DF.columns])
                msg = f"已删除 {len(cols)} 列"
            elif action == "to_numeric":
                for c in cols:
                    if c in DF.columns:
                        DF[c] = pd.to_numeric(DF[c], errors="coerce")
                msg = f"已将 {len(cols)} 列转为数值"
            elif action == "fill_empty":
                val = params.get("value", "")
                for c in cols:
                    if c in DF.columns:
                        DF[c] = DF[c].fillna("").replace("", val)
                msg = f"已填充 {len(cols)} 列的空值"
            elif action == "prefix_cols":
                prefix = params.get("text", "")
                for c in cols:
                    if c in DF.columns:
                        DF.rename(columns={c: prefix + str(c)}, inplace=True)
                msg = f"已为 {len(cols)} 列添加前缀"
            elif action == "suffix_cols":
                suffix = params.get("text", "")
                for c in cols:
                    if c in DF.columns:
                        DF.rename(columns={c: str(c) + suffix}, inplace=True)
                msg = f"已为 {len(cols)} 列添加后缀"
            elif action == "strip_all":
                for c in cols:
                    if c in DF.columns:
                        DF[c] = DF[c].astype(str).str.strip()
                msg = f"已去除 {len(cols)} 列的空格"
            elif action == "upper_all":
                for c in cols:
                    if c in DF.columns:
                        DF[c] = DF[c].astype(str).str.upper()
                msg = f"已将 {len(cols)} 列转大写"
            elif action == "lower_all":
                for c in cols:
                    if c in DF.columns:
                        DF[c] = DF[c].astype(str).str.lower()
                msg = f"已将 {len(cols)} 列转小写"
            else:
                ok, msg = False, f"未知批量操作: {action}"
        except Exception as e:
            ok, msg = False, f"批量操作出错: {str(e)}"

    return jsonify({"ok": True, "message": msg})


# ---------------------------------------------------------------------------
# 打印友好视图
# ---------------------------------------------------------------------------
@app.route("/print")
def print_view():
    with DATA_LOCK:
        df = DF.copy()
    cols = list(df.columns)
    rows = df.where(pd.notna(df), None).to_dict(orient="records")

    header_html = "".join("<th>" + c + "</th>" for c in cols)
    body_html = ""
    for i, row in enumerate(rows):
        body_html += "<tr>" + "".join("<td>" + str(row.get(c, "")) + "</td>" for c in cols) + "</tr>"

    return render_template_string("""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8"><title>ACM 数据打印</title>
<style>
  @media print {
    body { -webkit-print-color-adjust: exact; print-color-adjust: exact; }
    .no-print { display: none; }
  }
  body { font-family: "Microsoft YaHei",sans-serif; margin: 20px; }
  h1 { color: #333; font-size: 20px; margin-bottom: 4px; }
  .meta { color: #888; font-size: 11px; margin-bottom: 16px; }
  table { width: 100%; border-collapse: collapse; font-size: 11px; }
  th { background: #333; color: #fff; padding: 6px 10px; text-align: left; }
  td { padding: 5px 10px; border-bottom: 1px solid #ddd; }
  tr:nth-child(even) td { background: #f7f7f7; }
  .no-print { margin-bottom: 12px; }
  .no-print button { padding: 8px 16px; cursor: pointer; }
</style></head>
<body>
<div class="no-print">
  <button onclick="window.print()">🖨 打印</button>
  <button onclick="window.close()">关闭</button>
</div>
<h1>ACM 队伍数据</h1>
<div class="meta">""" + str(len(rows)) + " 行 × " + str(len(cols)) + " 列 | " + pd.Timestamp.now().strftime("%Y-%m-%d %H:%M") + """</div>
<table><thead><tr>""" + header_html + """</tr></thead>
<tbody>""" + body_html + """</tbody></table>
</body></html>""")


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import socket

    def _find_port(start=5000, end=5010):
        """自动查找可用端口 (兼容 macOS AirPlay 占用 5000)"""
        for p in range(start, end + 1):
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                if s.connect_ex(("127.0.0.1", p)) != 0:
                    return p
        return start  # fallback

    PORT = _find_port(5000, 5010)
    print("=" * 60)
    print(f"  ACM 队伍数据管理平台  —  http://127.0.0.1:{PORT}")
    print("  SQLite | 备份 | Diff | NL查询 | 打印 | 批量操作")
    if PORT != 5000:
        print(f"  (端口 5000 被占用，自动切换到 {PORT})")
    print("=" * 60)
    app.run(host="127.0.0.1", port=PORT, debug=False)
